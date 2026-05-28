from __future__ import annotations

import json
import math
import os
from datetime import date, datetime
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
SOURCE = Path(os.environ.get("REMUN_SOURCE", BASE_DIR / "data" / "detalle_remuneraciones.xlsx"))
OUTPUT = Path(os.environ.get("REMUN_OUTPUT", BASE_DIR / "index.html"))

HEADER_ROW_INDEX = 5

HABER_CONCEPT_RANGE = range(23, 86)  # 1-based Excel columns: Sueldo Base through Diferencia Colación.
LICENCIA_COLS = (13, 14, 15, 19)
AUSENTISMO_COLS = (13, 14, 15, 16, 17, 18, 19, 20, 21, 22)

MONTH_NAMES = {
    "01": "Ene",
    "02": "Feb",
    "03": "Mar",
    "04": "Abr",
    "05": "May",
    "06": "Jun",
    "07": "Jul",
    "08": "Ago",
    "09": "Sep",
    "10": "Oct",
    "11": "Nov",
    "12": "Dic",
}


def num(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return 0.0
        return float(value)
    try:
        return float(str(value).replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def clean(value) -> str:
    return str(value or "").strip()


def json_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def month_label(period: str) -> str:
    year, month = period.split("-")
    return f"{MONTH_NAMES.get(month, month)} {year}"


def pct(numerator: float, denominator: float) -> float:
    return round((numerator / denominator * 100), 1) if denominator else 0.0


def summarize(rows: list[dict]) -> dict:
    dot = len(rows)
    dias = sum(r["dias"] for r in rows)
    sueldo_base = sum(r["sueldo_base"] for r in rows)
    total_haberes = sum(r["total_haberes"] for r in rows)
    liquido = sum(r["sueldo_liquido"] for r in rows)
    descuentos = sum(r["total_descuentos"] for r in rows)
    hhee = sum(r["hhee"] for r in rows)
    licencia_dias = sum(r["licencia_dias"] for r in rows)
    ausentismo_dias = sum(r["ausentismo_dias"] for r in rows)
    grat = sum(r["gratificacion"] for r in rows)
    movilizacion = sum(r["movilizacion"] for r in rows)
    colacion = sum(r["colacion"] for r in rows)
    cero = sum(1 for r in rows if r["dias"] == 0)
    parcial = sum(1 for r in rows if 0 < r["dias"] < 30)
    return {
        "dotacion": dot,
        "fte": round(sum(r["dias"] / 30 for r in rows), 2),
        "dias_prom": round(dias / dot, 1) if dot else 0,
        "cero_dias": cero,
        "menos30": parcial,
        "rotacion_pct": pct(cero + parcial, dot),
        "ausentes_pct": pct(cero, dot),
        "total_haberes": round(total_haberes),
        "sueldo_base": round(sueldo_base),
        "sueldo_liquido": round(liquido),
        "total_descuentos": round(descuentos),
        "gratificacion": round(grat),
        "hhee": round(hhee),
        "licencia_dias": round(licencia_dias, 1),
        "ausentismo_dias": round(ausentismo_dias, 1),
        "licencia_personas": sum(1 for r in rows if r["licencia_dias"] > 0),
        "ausentismo_personas": sum(1 for r in rows if r["ausentismo_dias"] > 0),
        "movilizacion": round(movilizacion),
        "colacion": round(colacion),
        "sb_promedio": round(sueldo_base / dot) if dot else 0,
        "liq_promedio": round(liquido / dot) if dot else 0,
        "pct_fte": pct(sum(r["dias"] / 30 for r in rows), dot),
    }


def group_rows(rows: list[dict], keys: tuple[str, ...]) -> list[dict]:
    buckets: dict[tuple[str, ...], list[dict]] = defaultdict(list)
    for row in rows:
        buckets[tuple(row[k] for k in keys)].append(row)

    out = []
    for key, items in buckets.items():
        item = summarize(items)
        for name, value in zip(keys, key):
            item[name] = value
        item["key"] = " · ".join(k or "Sin dato" for k in key)
        out.append(item)
    return sorted(out, key=lambda r: r["total_haberes"], reverse=True)


def group_concepts(rows: list[dict], concept_cols: list[dict]) -> dict:
    grouped = {
        "total": Counter(),
        "empresa": defaultdict(Counter),
        "sede": defaultdict(Counter),
        "empresa_sede": defaultdict(Counter),
    }
    for row in rows:
        empresa = row["_raw_empresa"]
        sede = row["_raw_sede"]
        empresa_sede = f"{empresa} · {sede}"
        for col in concept_cols:
            idx, concept = col["idx"], col["label"]
            value = num(row["_raw"][idx] if idx < len(row["_raw"]) else 0)
            grouped["total"][concept] += value
            grouped["empresa"][empresa][concept] += value
            grouped["sede"][sede][concept] += value
            grouped["empresa_sede"][empresa_sede][concept] += value
    return {
        "total": dict(grouped["total"]),
        "empresa": {k: dict(v) for k, v in grouped["empresa"].items()},
        "sede": {k: dict(v) for k, v in grouped["sede"].items()},
        "empresa_sede": {k: dict(v) for k, v in grouped["empresa_sede"].items()},
    }


def build_data() -> dict:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Detalle"]
    rows_iter = ws.iter_rows(values_only=True)

    for _ in range(HEADER_ROW_INDEX - 1):
        next(rows_iter)
    headers = list(next(rows_iter))
    header_index = {h: i for i, h in enumerate(headers) if h}

    concept_cols = []
    for one_based in HABER_CONCEPT_RANGE:
        idx = one_based - 1
        label = clean(headers[idx])
        if label and label not in {"Suma Haberes", "Sueldo Líquido"}:
            concept_cols.append({"idx": idx, "label": label, "type": "haber"})
    discount_start = header_index.get("Cotizacion AFP")
    discount_end = min(
        [i for name in ("Aporte a CCAF", "Mutual", "Sueldo Líquido") if (i := header_index.get(name)) is not None],
        default=len(headers),
    )
    if discount_start is not None:
        for idx in range(discount_start, discount_end):
            label = clean(headers[idx])
            if label and label not in {"Total Rebajas"}:
                concept_cols.append({"idx": idx, "label": label, "type": "descuento"})

    by_month_rows: dict[str, list[dict]] = defaultdict(list)
    concept_sums: dict[str, Counter] = defaultdict(Counter)
    companies = Counter()

    for raw in rows_iter:
        if not raw or not raw[0] or raw[0] == "Empresa":
            continue
        period = clean(raw[header_index["Proceso"]])
        empresa = clean(raw[header_index["Nombre empresa"]])
        sede = clean(raw[header_index["Sede"]]) or "Sin sede"
        row = {
            "_raw": raw,
            "_raw_empresa": empresa,
            "_raw_sede": sede,
            "proceso": period,
            "empresa": empresa,
            "rut_empresa": clean(raw[header_index["Rut empresa"]]),
            "sede": sede,
            "nombre": clean(raw[header_index["Nombre"]]),
            "rut": clean(raw[header_index["Rut"]]),
            "contrato": clean(raw[header_index["Contrato"]]),
            "cargo": clean(raw[header_index["Cargo"]]),
            "dias": num(raw[header_index["Días Trabajados"]]),
            "sueldo_base": num(raw[header_index["Sueldo Base"]]),
            "total_haberes": num(raw[header_index["Suma Haberes"]]),
            "sueldo_liquido": num(raw[header_index["Sueldo Líquido"]]),
            "total_descuentos": num(raw[header_index["Total Rebajas"]]),
            "gratificacion": num(raw[header_index["Gratificación"]]),
            "hhee": num(raw[header_index["Horas Extras Empresa 50%"]]),
            "licencia_dias": sum(num(raw[i - 1] if i - 1 < len(raw) else 0) for i in LICENCIA_COLS),
            "ausentismo_dias": sum(num(raw[i - 1] if i - 1 < len(raw) else 0) for i in AUSENTISMO_COLS),
            "movilizacion": num(raw[header_index["Movilizacion"]]),
            "colacion": num(raw[header_index["Colacion"]]),
            "bono_manip_pae": num(raw[header_index["Bono Manipuladora Pae"]]) + num(raw[header_index["Bono Manipuladora Pae I"]]),
            "concept_values": {
                col["label"]: value
                for col in concept_cols
                for idx in [col["idx"]]
                if (value := num(raw[idx] if idx < len(raw) else 0)) != 0
            },
        }
        by_month_rows[period].append(row)
        companies[empresa] += 1
        for col in concept_cols:
            idx, concept = col["idx"], col["label"]
            concept_sums[period][concept] += num(raw[idx] if idx < len(raw) else 0)

    months = sorted(by_month_rows)
    data = {
        "metadata": {
            "source": SOURCE.name,
            "generated_from": "Detalle",
            "month_count": len(months),
            "record_count": sum(len(v) for v in by_month_rows.values()),
            "company_count": len(companies),
            "raw_headers": [json_value(h) for h in headers],
        },
        "months": [{"id": m, "label": month_label(m)} for m in months],
        "kpis": {},
        "groups": {"empresa": {}, "sede": {}, "empresa_sede": {}},
        "concepts": {},
        "concept_groups": {},
        "concept_options": sorted({col["label"] for col in concept_cols}),
        "concept_types": {col["label"]: col["type"] for col in concept_cols},
        "details": {},
    }

    for month in months:
        rows = by_month_rows[month]
        data["kpis"][month] = summarize(rows)
        data["groups"]["empresa"][month] = group_rows(rows, ("empresa",))
        data["groups"]["sede"][month] = group_rows(rows, ("sede",))
        data["groups"]["empresa_sede"][month] = group_rows(rows, ("empresa", "sede"))
        data["concepts"][month] = dict(sorted(concept_sums[month].items()))
        data["concept_groups"][month] = group_concepts(rows, concept_cols)
        data["details"][month] = [
            {
                "rut": r["rut"],
                "nombre": r["nombre"],
                "cargo": r["cargo"],
                "contrato": r["contrato"],
                "empresa": r["empresa"],
                "sede": r["sede"],
                "dias": r["dias"],
                "fte": round(r["dias"] / 30, 2),
                "sueldo_base": round(r["sueldo_base"]),
                "total_haberes": round(r["total_haberes"]),
                "total_descuentos": round(r["total_descuentos"]),
                "sueldo_liquido": round(r["sueldo_liquido"]),
                "hhee": round(r["hhee"]),
                "licencia_dias": round(r["licencia_dias"], 1),
                "ausentismo_dias": round(r["ausentismo_dias"], 1),
                "bono_manip_pae": round(r["bono_manip_pae"]),
                "raw": [json_value(v) for v in r["_raw"]],
                "concepts": {k: round(v) for k, v in r["concept_values"].items()},
            }
            for r in rows
        ]

    return data


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Comparador Remuneraciones · 13 meses</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;600&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/xlsx-js-style@1.2.0/dist/xlsx.bundle.js"></script>
<style>
:root{
  --bg:#f4f5f9;--surface:#fff;--surface2:#f0f2f8;--border:#e2e6ef;--border2:#c9cedf;
  --text:#1a1d2e;--text2:#5a6080;--text3:#9198b5;
  --blue:#2563eb;--blue-lt:#eff4ff;--blue-mid:#93b4f8;
  --green:#059669;--green-lt:#ecfdf5;--red:#dc2626;--red-lt:#fef2f2;
  --amber:#d97706;--amber-lt:#fffbeb;--purple:#7c3aed;--purple-lt:#f5f3ff;
  --sh0:0 1px 3px rgba(0,0,0,.06),0 1px 2px rgba(0,0,0,.04);
  --sh1:0 4px 14px rgba(0,0,0,.08);--sh2:0 10px 32px rgba(0,0,0,.10);
  --r:10px;--f:'Inter',system-ui,sans-serif;--m:'IBM Plex Mono',monospace;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{background:var(--bg);color:var(--text);font-family:var(--f);font-size:13.5px;line-height:1.55}
.header{background:var(--surface);border-bottom:1px solid var(--border);padding:0 28px;display:flex;align-items:center;position:sticky;top:0;z-index:200;box-shadow:var(--sh0);gap:12px;flex-wrap:wrap}
.h-brand{display:flex;align-items:center;gap:10px;padding:13px 18px 13px 0;border-right:1px solid var(--border)}
.h-icon{width:34px;height:34px;border-radius:8px;background:linear-gradient(135deg,#2563eb,#1d4ed8);display:flex;align-items:center;justify-content:center;color:#fff;font-size:13px;font-weight:700;font-family:var(--m)}
.h-name{font-size:14px;font-weight:700;line-height:1.2}.h-sub{font-size:11px;color:var(--text3)}
.chip{display:inline-flex;align-items:center;font-size:11.5px;font-weight:600;padding:3px 10px;border-radius:20px}.chip-b{background:var(--blue-lt);color:var(--blue)}.chip-g{background:var(--green-lt);color:var(--green)}.chip-vs{color:var(--text3);font-size:11px}
.h-actions{margin-left:auto;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.btn{display:inline-flex;align-items:center;gap:5px;padding:6px 13px;border-radius:7px;border:1px solid var(--border2);background:var(--surface);color:var(--text2);font-family:var(--f);font-size:12.5px;font-weight:500;cursor:pointer;transition:all .15s}
.btn:hover{border-color:var(--blue);color:var(--blue);background:var(--blue-lt)}
.nav{background:var(--surface);border-bottom:1px solid var(--border);padding:0 28px;display:flex;overflow-x:auto}.nav::-webkit-scrollbar{display:none}
.tab{padding:12px 16px;font-size:12.5px;font-weight:500;color:var(--text3);cursor:pointer;border-bottom:2px solid transparent;transition:all .15s;white-space:nowrap}.tab:hover{color:var(--text2)}.tab.active{color:var(--blue);border-bottom-color:var(--blue);font-weight:600}
.main{padding:22px 28px;max-width:1440px;margin:0 auto}.sec{display:none}.sec.active{display:block}
.sec-hd{margin-bottom:18px}.sec-title{font-size:16px;font-weight:700}.sec-sub{font-size:12px;color:var(--text3);margin-top:2px}
.controlbar{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);box-shadow:var(--sh0);padding:14px 16px;margin-bottom:18px;display:flex;gap:12px;align-items:end;flex-wrap:wrap}
.ctrl{display:flex;flex-direction:column;gap:4px}.ctrl label{font-size:10.5px;text-transform:uppercase;letter-spacing:.5px;color:var(--text3);font-weight:700}
.fsel,.srch,.file-input{background:var(--surface);border:1px solid var(--border2);color:var(--text);padding:7px 10px;border-radius:7px;font-family:var(--f);font-size:12.5px;outline:none}.srch{min-width:240px}.file-input{max-width:260px}.fsel:focus,.srch:focus,.file-input:focus{border-color:var(--blue);box-shadow:0 0 0 3px rgba(37,99,235,.1)}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(165px,1fr));gap:11px;margin-bottom:22px}
.kpi{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);padding:15px 16px;box-shadow:var(--sh0);transition:box-shadow .15s,transform .15s;border-top:3px solid var(--blue)}
.kpi:hover{box-shadow:var(--sh1);transform:translateY(-1px)}.kpi-lbl{font-size:10.5px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px}.kpi-val{font-size:21px;font-weight:700;line-height:1;margin-bottom:5px;font-family:var(--m)}
.kpi-foot{display:flex;align-items:center;justify-content:space-between;font-size:11px}.kpi-prev{color:var(--text3)}.kpi-dif{font-weight:600;font-family:var(--m);font-size:11px}.dp{color:var(--green)}.dn{color:var(--red)}.d0{color:var(--text3)}
.panel{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);box-shadow:var(--sh0);margin-bottom:18px;overflow:hidden}.panel-hd{padding:13px 18px;border-bottom:1px solid var(--border);display:flex;align-items:flex-start;justify-content:space-between;gap:10px}.panel-title{font-size:13px;font-weight:700}.panel-sub{font-size:11.5px;color:var(--text3);margin-top:2px}.panel-body{padding:18px}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px}.g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin-bottom:18px}@media(max-width:960px){.g2,.g3{grid-template-columns:1fr}.h-actions{margin-left:0}}
.bc{display:flex;flex-direction:column;gap:8px}.br{display:flex;align-items:center;gap:8px}.bl{width:190px;font-size:12px;color:var(--text2);text-align:right;flex-shrink:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.bt{flex:1;height:22px;background:var(--surface2);border-radius:5px;overflow:hidden}.bf{height:100%;border-radius:5px;display:flex;align-items:center;padding:0 8px;min-width:2px}.bf span{font-size:10.5px;font-weight:600;color:#fff;white-space:nowrap;font-family:var(--m)}.ba{width:80px;font-size:11px;color:var(--text3);font-family:var(--m)}
.tw{overflow:auto}.tw::-webkit-scrollbar{height:4px;width:4px}.tw::-webkit-scrollbar-track{background:var(--surface2)}.tw::-webkit-scrollbar-thumb{background:var(--border2);border-radius:2px}.sy{max-height:520px;overflow-y:auto}
table{width:100%;border-collapse:collapse;font-size:12.5px}thead th{background:var(--surface2);color:var(--text2);font-size:10.5px;font-weight:600;letter-spacing:.4px;text-transform:uppercase;padding:9px 11px;text-align:left;position:sticky;top:0;border-bottom:1px solid var(--border);white-space:nowrap}tbody tr{border-bottom:1px solid var(--border);transition:background .1s}tbody tr:last-child{border-bottom:none}tbody tr:hover{background:var(--blue-lt)}tbody td{padding:8px 11px;vertical-align:middle}.nr{text-align:right;font-family:var(--m);font-size:12px}.tn{font-weight:500}
.sec-wide{width:calc(100vw - 28px);margin-left:calc(50% - 50vw + 14px);margin-right:calc(50% - 50vw + 14px)}
.top-scroll{height:18px;overflow-x:auto;overflow-y:hidden;border-bottom:1px solid var(--border);background:var(--surface2)}.top-scroll div{height:1px}.top-scroll::-webkit-scrollbar{height:14px}.top-scroll::-webkit-scrollbar-track{background:#e7ecf5}.top-scroll::-webkit-scrollbar-thumb{background:#9aa8bd;border-radius:8px;border:3px solid #e7ecf5}.book-wrap{max-height:calc(100vh - 270px);min-height:520px;overflow:auto}.book-wrap::-webkit-scrollbar{height:14px;width:8px}.book-wrap::-webkit-scrollbar-track{background:#e7ecf5}.book-wrap::-webkit-scrollbar-thumb{background:#9aa8bd;border-radius:8px;border:3px solid #e7ecf5}.book-table{min-width:6200px;border-collapse:separate;border-spacing:0}.book-table th,.book-table td{height:34px;border-bottom:1px solid var(--border);background:var(--surface);white-space:nowrap}.book-table thead th{z-index:20;background:var(--surface2);top:0;vertical-align:top;height:62px;padding:7px 8px}.book-table tbody tr:hover td{background:var(--blue-lt)}.book-table .fix{position:sticky;z-index:15;box-shadow:1px 0 0 var(--border)}.book-table thead .fix{z-index:30}.book-table .f1{left:0;width:220px;min-width:220px;max-width:220px}.book-table .f2{left:220px;width:180px;min-width:180px;max-width:180px}.book-table .f3{left:400px;width:270px;min-width:270px;max-width:270px}.book-table .fix{background:var(--surface)}.book-table thead .fix{background:var(--surface2)}.book-table .money-col{min-width:118px}.book-table .text-col{min-width:180px}.book-table .muted-cell{color:var(--text3);font-size:11.5px}.book-table .zero-cell{color:#c8cfdd}.th-label{display:block;margin-bottom:5px;overflow:hidden;text-overflow:ellipsis}.head-filter-btn{width:100%;height:24px;display:flex;align-items:center;justify-content:space-between;gap:5px;border:1px solid var(--border2);border-radius:4px;background:#fff;color:var(--text2);font-family:var(--f);font-size:10.5px;font-weight:600;letter-spacing:0;text-transform:none;padding:2px 6px;cursor:pointer}.head-filter-btn.active{border-color:var(--blue);background:var(--blue-lt);color:var(--blue)}.head-filter-btn span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.detail-filter-menu{position:fixed;z-index:1000;width:300px;max-height:390px;background:var(--surface);border:1px solid var(--border2);border-radius:8px;box-shadow:var(--sh2);padding:10px;display:none}.detail-filter-menu.open{display:block}.df-title{font-weight:800;font-size:12px;margin-bottom:8px}.df-search{width:100%;border:1px solid var(--border2);border-radius:6px;padding:7px 8px;font-family:var(--f);font-size:12px;margin-bottom:8px}.df-actions{display:flex;gap:6px;margin-bottom:8px}.df-actions button,.df-apply{border:1px solid var(--border2);background:var(--surface);border-radius:6px;padding:5px 8px;font-family:var(--f);font-size:11.5px;font-weight:700;cursor:pointer}.df-actions button:hover,.df-apply:hover{border-color:var(--blue);color:var(--blue);background:var(--blue-lt)}.df-list{max-height:235px;overflow:auto;border:1px solid var(--border);border-radius:6px}.df-opt{display:flex;align-items:center;gap:7px;padding:6px 8px;border-bottom:1px solid var(--border);font-size:12px;text-transform:none;letter-spacing:0}.df-opt:last-child{border-bottom:none}.df-opt input{margin:0}.df-opt span{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.df-note{font-size:11px;color:var(--text3);margin:7px 0}.df-foot{display:flex;justify-content:flex-end;margin-top:8px}
.bdg{display:inline-flex;align-items:center;padding:2px 7px;border-radius:4px;font-size:10.5px;font-weight:600;font-family:var(--m)}.bg{background:var(--green-lt);color:var(--green)}.br2{background:var(--red-lt);color:var(--red)}.by{background:var(--amber-lt);color:var(--amber)}.bb{background:var(--blue-lt);color:var(--blue)}.bgr{background:var(--surface2);color:var(--text2)}
.cmp-banner{background:linear-gradient(135deg,var(--blue) 0%,#1d4ed8 100%);border-radius:var(--r);padding:18px 26px;margin-bottom:18px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:14px;color:#fff}.cbl{font-size:11px;font-weight:600;opacity:.75;text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px}.cbv{font-family:var(--m);font-size:26px;font-weight:700;line-height:1}.cbs{font-size:12px;font-weight:600;opacity:.8;margin-top:3px}
.cc-hd,.cc-row{display:grid;grid-template-columns:220px 1fr 1fr 120px 150px;gap:10px;align-items:center}.cc-hd{padding:8px 14px;background:var(--surface2);border-bottom:1px solid var(--border);font-size:10.5px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:.4px}.cc-section{display:flex;align-items:center;justify-content:space-between;gap:10px;padding:10px 14px;background:#eef4ff;border-top:1px solid var(--border);border-bottom:1px solid var(--border);font-size:11px;font-weight:800;color:var(--blue);text-transform:uppercase;letter-spacing:.4px}.cc-section span{font-family:var(--m);font-size:11px;color:var(--text3);text-transform:none;letter-spacing:0}.cc-row{padding:9px 14px;border-bottom:1px solid var(--border);font-size:12.5px}.cc-row:hover{background:var(--surface2)}.cc-name{font-weight:600}.cc-num{font-family:var(--m);font-size:12px;text-align:right}.cc-bar{display:flex;align-items:center;gap:5px}.cc-mb{flex:1;height:5px;background:var(--surface2);border-radius:3px;overflow:hidden}.cc-mf{height:100%;border-radius:3px}
.rg{display:grid;grid-template-columns:repeat(3,1fr);gap:11px;margin-bottom:18px}.rc{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);padding:15px 18px;box-shadow:var(--sh0)}.rc.active{box-shadow:0 0 0 2px rgba(37,99,235,.18);background:var(--blue-lt)}.rv{font-family:var(--m);font-size:28px;font-weight:700;line-height:1;margin-bottom:3px}.rl{font-size:11.5px;color:var(--text3)}.rs{font-size:11px;color:var(--text3);margin-top:2px}
.hi-grid{display:grid;grid-template-columns:1.1fr .9fr;gap:14px;margin-bottom:18px}.hi-list{display:flex;flex-direction:column;gap:8px}.hi{display:flex;align-items:center;justify-content:space-between;gap:14px;padding:10px 12px;border:1px solid var(--border);border-radius:8px;background:var(--surface)}.hi-main{min-width:0}.hi-title{font-weight:700;font-size:12.5px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.hi-sub{font-size:11px;color:var(--text3);margin-top:1px}.hi-val{font-family:var(--m);font-weight:700;font-size:13px;white-space:nowrap}.tot-table td{padding:7px 10px}.dev-pill{display:inline-flex;align-items:center;border-radius:999px;padding:2px 8px;font-size:10.5px;font-weight:700;font-family:var(--m)}@media(max-width:960px){.hi-grid{grid-template-columns:1fr}}
.clickable{cursor:pointer}.clickable:hover{background:var(--blue-lt)}.selected-line{background:var(--blue-lt)}
.exp-btn{width:24px;height:24px;border:1px solid var(--border2);border-radius:6px;background:var(--surface);color:var(--text2);font-family:var(--m);font-weight:700;cursor:pointer}.exp-btn:hover{border-color:var(--blue);color:var(--blue);background:var(--blue-lt)}.sede-row td{background:#fafbff}.sede-name{padding-left:34px;color:var(--text2)}
.chip-x{display:inline-flex;align-items:center;gap:6px;border:1px solid var(--border2);background:var(--surface);border-radius:999px;padding:4px 9px;font-size:11.5px;font-weight:600;color:var(--text2)}.chip-x button{border:none;background:transparent;color:var(--red);font-weight:800;cursor:pointer;font-size:13px;line-height:1}.chip-wrap{display:flex;gap:7px;flex-wrap:wrap;align-items:center}
canvas{max-width:100%}
::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:var(--bg)}::-webkit-scrollbar-thumb{background:var(--border2);border-radius:2px}
</style>
</head>
<body>
<header class="header">
  <div class="h-brand">
    <div class="h-icon">13M</div>
    <div><div class="h-name">Remuneraciones</div><div class="h-sub">Empresa / Sede</div></div>
  </div>
  <span class="chip chip-b" id="chipBase"></span><span class="chip-vs">vs</span><span class="chip chip-g" id="chipComp"></span>
  <span style="font-size:11.5px;color:var(--text3)" id="sourceBadge"></span>
  <div class="h-actions"><button class="btn" onclick="window.print()">Imprimir</button></div>
</header>
<nav class="nav" id="navTabs">
  <div class="tab active" data-sec="dashboard">Dashboard</div>
  <div class="tab" data-sec="kpis">KPIs</div>
  <div class="tab" data-sec="conceptos">Diferencia mensual</div>
  <div class="tab" data-sec="auditoria">Auditoría</div>
  <div class="tab" data-sec="descargas">Descargas</div>
  <div class="tab" data-sec="detalle">Detalle</div>
  <div class="tab" data-sec="beta">Beta</div>
</nav>
<main class="main">
  <div class="controlbar">
    <div class="ctrl"><label>Mes base</label><select class="fsel" id="selBase"></select></div>
    <div class="ctrl"><label>Mes comparación</label><select class="fsel" id="selComp"></select></div>
    <div class="ctrl"><label>Empresa</label><select class="fsel" id="selEmpresa"><option value="">Todas las empresas</option></select></div>
    <div class="ctrl"><label>Sede</label><select class="fsel" id="selSede"><option value="">Todas las sedes</option></select></div>
    <div class="ctrl"><label>Cargar archivo</label><input class="file-input" id="fileUpload" type="file" accept=".xlsx,.xls" onchange="handleFileUpload(event)"></div>
  </div>

  <section class="sec active" id="sec-dashboard">
    <div class="sec-hd"><div class="sec-title">Dashboard General</div><div class="sec-sub" id="dashSub"></div></div>
    <div class="panel">
      <div class="panel-hd"><div><div class="panel-title">División por Empresa</div><div class="panel-sub">Expande una empresa para ver sus sedes; click en una fila abre el detalle</div></div></div>
      <div class="tw sy"><table><thead><tr><th></th><th>Empresa / Sede</th><th class="nr">Dotación</th><th class="nr">FTE</th><th class="nr">Suma Haberes</th><th class="nr">Sueldo Líquido</th><th class="nr">Sueldo Base</th><th class="nr">Δ Haberes</th></tr></thead><tbody id="companyDivisionBody"></tbody></table></div>
    </div>
    <div class="kpi-grid" id="kpiGrid"></div>
    <div class="hi-grid">
      <div class="panel"><div class="panel-hd"><div><div class="panel-title">Totales del período</div><div class="panel-sub" id="totalsSub"></div></div></div><div class="tw"><table class="tot-table"><tbody id="totalsBody"></tbody></table></div></div>
      <div class="panel"><div class="panel-hd"><div><div class="panel-title">Highlights</div><div class="panel-sub">Principales señales del comparativo</div></div></div><div class="panel-body"><div class="hi-list" id="highlights"></div></div></div>
    </div>
  </section>

  <section class="sec" id="sec-kpis">
    <div class="sec-hd"><div class="sec-title">KPIs Operacionales</div><div class="sec-sub" id="kpiOpsSub">Rotación, ausentismo y licencias con jerarquía Empresa / Sede / Trabajador</div></div>
    <div class="rg" id="kpiOpsGrid"></div>
    <div class="panel">
      <div class="panel-hd"><div><div class="panel-title">Rotación mensual</div><div class="panel-sub">Ingresos y egresos entre mes base y mes comparación</div></div></div>
      <div class="tw sy"><table><thead><tr><th></th><th>Empresa / Sede / Trabajador</th><th>Tipo</th><th class="nr">Días</th><th class="nr">FTE</th><th class="nr">Suma Haberes</th></tr></thead><tbody id="kpiRotBody"></tbody></table></div>
    </div>
    <div class="g2">
      <div class="panel">
        <div class="panel-hd"><div><div class="panel-title">Ausentismo</div><div class="panel-sub">Licencias, permisos y faltas del mes comparación</div></div></div>
        <div class="tw sy"><table><thead><tr><th></th><th>Empresa / Sede / Trabajador</th><th class="nr">Días ausentismo</th><th class="nr">Días trabajados</th><th class="nr">FTE</th></tr></thead><tbody id="kpiAusBody"></tbody></table></div>
      </div>
      <div class="panel">
        <div class="panel-hd"><div><div class="panel-title">Licencias</div><div class="panel-sub">Días de licencia médica y accidente</div></div></div>
        <div class="tw sy"><table><thead><tr><th></th><th>Empresa / Sede / Trabajador</th><th class="nr">Días licencia</th><th class="nr">Días trabajados</th><th class="nr">FTE</th></tr></thead><tbody id="kpiLicBody"></tbody></table></div>
      </div>
    </div>
  </section>

  <section class="sec" id="sec-conceptos">
    <div class="sec-hd"><div class="sec-title">Diferencia mensual</div><div class="sec-sub" id="conceptSub">Haz click en una línea para ver el monto por sede</div></div>
    <div class="controlbar">
      <div class="ctrl"><label>Cantidad</label><select class="fsel" id="conceptLimit"><option value="15">Top 15</option><option value="25">Top 25</option><option value="40">Top 40</option><option value="all">Todos</option></select></div>
      <div class="ctrl"><label>Orden</label><select class="fsel" id="conceptSort"><option value="abs">Magnitud absoluta</option><option value="up">Mayores alzas</option><option value="down">Mayores bajas</option><option value="name">Nombre</option></select></div>
      <div class="ctrl"><label>Buscar concepto</label><input class="srch" id="conceptSearch" placeholder="Ej. bono, gratificación, movilización"></div>
    </div>
    <div class="panel">
      <div class="cc-hd"><div>Concepto</div><div style="text-align:right" id="ccBase"></div><div style="text-align:right" id="ccComp"></div><div style="text-align:right">Diferencia</div><div>Magnitud</div></div>
      <div id="ccRows"></div><div style="padding:9px 14px;background:var(--surface2);border-top:2px solid var(--border2)" id="ccTot"></div>
    </div>
  </section>

  <section class="sec" id="sec-auditoria">
    <div class="sec-hd"><div class="sec-title">Auditoría</div><div class="sec-sub" id="auditSub">Controles sobre el mes comparación y filtros Empresa/Sede</div></div>
    <div class="rg" id="auditGrid"></div>
    <div class="panel">
      <div class="panel-hd"><div><div class="panel-title">Personas con 0 días y haberes distintos de cero</div><div class="panel-sub">Click en una fila abre el detalle filtrado por sede</div></div></div>
      <div class="tw sy"><table><thead><tr><th>Empresa</th><th>Sede</th><th>RUT</th><th>Nombre</th><th>Cargo</th><th class="nr">Días</th><th class="nr">Suma Haberes</th><th class="nr">Líquido</th></tr></thead><tbody id="auditZeroBody"></tbody></table></div>
    </div>
    <div class="g2">
      <div class="panel">
        <div class="panel-hd"><div><div class="panel-title">Manipulador (a) PAE con sueldo base $539.000 o proporcional</div><div class="panel-sub">Esperado = 539.000 × días / 30</div></div></div>
        <div class="tw sy"><table><thead><tr><th>Empresa</th><th>Sede</th><th class="nr">Total</th><th class="nr">OK</th><th class="nr">Diferencias</th><th class="nr">% OK</th></tr></thead><tbody id="auditBaseBody"></tbody><tfoot id="auditBaseFoot" style="background:var(--surface2);font-weight:700;border-top:2px solid var(--border2)"></tfoot></table></div>
      </div>
      <div class="panel">
        <div class="panel-hd"><div><div class="panel-title">Manipulador (a) PAE con Bono Manipuladora PAE</div><div class="panel-sub">Monto agregado por sede</div></div></div>
        <div class="tw sy"><table><thead><tr><th>Empresa</th><th>Sede</th><th class="nr">Personas</th><th class="nr">Con bono</th><th class="nr">% con bono</th><th class="nr">Monto bono</th></tr></thead><tbody id="auditBonoBody"></tbody><tfoot id="auditBonoFoot" style="background:var(--surface2);font-weight:700;border-top:2px solid var(--border2)"></tfoot></table></div>
      </div>
    </div>
  </section>

  <section class="sec" id="sec-descargas">
    <div class="sec-hd"><div class="sec-title">Descargas</div><div class="sec-sub" id="downloadSub">Genera archivos de revisión usando siempre los últimos 2 meses del archivo cargado</div></div>
    <div class="rg" id="downloadGrid"></div>
    <div class="panel">
      <div class="panel-hd"><div><div class="panel-title">Archivos disponibles</div><div class="panel-sub">Cada descarga incluye Resumen por empresa, Resumen completo, Detalle individual, Mes anterior y Mes actual</div></div></div>
      <div class="tw"><table><thead><tr><th>Archivo</th><th>Regla aplicada</th><th class="nr">Base</th><th class="nr">Actual</th><th class="nr">Acción</th></tr></thead><tbody id="downloadBody"></tbody></table></div>
    </div>
  </section>

  <section class="sec sec-wide" id="sec-detalle">
    <div class="sec-hd"><div class="sec-title">Libro de Remuneraciones</div><div class="sec-sub">Vista detallada a pantalla completa, con cabecera fija, filtros por columna y columnas congeladas hasta Nombre</div></div>
    <div class="controlbar">
      <input class="srch" id="srch" placeholder="Buscar nombre, RUT, cargo, empresa o sede..." oninput="renderDetail()">
      <select class="fsel" id="detMonth" onchange="refreshDetailFilters();renderDetail()"></select>
      <select class="fsel" id="detEmpresa" onchange="refreshDetailFilters();renderDetail()"><option value="">Todas las empresas</option></select>
      <select class="fsel" id="detSede" onchange="refreshDetailFilters();renderDetail()"><option value="">Todas las sedes</option></select>
      <select class="fsel" id="detCargo" onchange="renderDetail()"><option value="">Todos los cargos</option></select>
      <select class="fsel" id="detDias" onchange="renderDetail()"><option value="">Todos los días</option><option value="0">0 días</option><option value="p">Parciales (&lt;30)</option><option value="f">30 días</option></select>
      <select class="fsel" id="detConcept" onchange="renderDetail()"><option value="">Todos los conceptos</option></select>
      <span style="font-size:11px;color:var(--text3);margin-left:auto" id="detCnt"></span>
    </div>
    <div class="panel"><div class="top-scroll" id="detailTopScroll"><div id="detailTopInner"></div></div><div class="book-wrap" id="detailBookWrap"><table class="book-table"><thead><tr id="detHead"></tr></thead><tbody id="detBody"></tbody></table></div></div>
    <div class="detail-filter-menu" id="detailFilterMenu"></div>
  </section>

  <section class="sec" id="sec-beta">
    <div class="sec-hd"><div class="sec-title">Beta</div><div class="sec-sub">Comparativo mensual y análisis de desviaciones</div></div>
    <div class="panel-hd"><div><div class="panel-title">Comparativo Mensual</div><div class="panel-sub">Variación por Empresa/Sede y conceptos seleccionados</div></div></div>
    <div class="controlbar">
      <div class="ctrl"><label>Agregar columna</label><select class="fsel" id="cmpMetricAdd"></select></div>
      <button class="btn" type="button" onclick="addCompareMetric()">Agregar concepto</button>
      <div class="chip-wrap" id="cmpMetricChips"></div>
    </div>
    <div class="cmp-banner" id="cmpBanner"></div>
    <div class="panel"><div class="tw sy"><table><thead><tr id="cmpHead"></tr></thead><tbody id="cmpBody"></tbody><tfoot id="cmpFoot" style="background:var(--surface2);font-weight:700;border-top:2px solid var(--border2)"></tfoot></table></div></div>

    <div class="sec-hd" style="margin-top:26px"><div class="sec-title">Análisis de Desviaciones</div><div class="sec-sub">Brechas por Empresa/Sede, conceptos y ratios operacionales</div></div>
    <div class="rg" id="devGrid"></div>
    <div class="g2">
      <div class="panel"><div class="panel-hd"><div><div class="panel-title">Mayores alzas por Empresa/Sede</div><div class="panel-sub" id="devUpSub"></div></div></div><div class="tw sy"><table><thead><tr><th>Empresa/Sede</th><th class="nr">Base</th><th class="nr">Comparación</th><th class="nr">Δ</th><th class="nr">Δ%</th><th class="nr">Δ Dot.</th></tr></thead><tbody id="devUpBody"></tbody></table></div></div>
      <div class="panel"><div class="panel-hd"><div><div class="panel-title">Mayores bajas por Empresa/Sede</div><div class="panel-sub" id="devDownSub"></div></div></div><div class="tw sy"><table><thead><tr><th>Empresa/Sede</th><th class="nr">Base</th><th class="nr">Comparación</th><th class="nr">Δ</th><th class="nr">Δ%</th><th class="nr">Δ Dot.</th></tr></thead><tbody id="devDownBody"></tbody></table></div></div>
    </div>
    <div class="g2">
      <div class="panel"><div class="panel-hd"><div><div class="panel-title">Desviaciones por concepto</div><div class="panel-sub">Top drivers positivos y negativos</div></div></div><div class="tw sy"><table><thead><tr><th>Concepto</th><th class="nr">Base</th><th class="nr">Comparación</th><th class="nr">Δ</th><th class="nr">Δ%</th></tr></thead><tbody id="devConceptBody"></tbody></table></div></div>
      <div class="panel"><div class="panel-hd"><div><div class="panel-title">Ratios críticos</div><div class="panel-sub">Cambios en costo promedio, FTE y ausencias</div></div></div><div class="tw"><table><thead><tr><th>Ratio</th><th class="nr">Base</th><th class="nr">Comparación</th><th class="nr">Δ</th></tr></thead><tbody id="devRatioBody"></tbody></table></div></div>
    </div>
  </section>
</main>
<script>
let DATA = __DATA__;
const state = {base: DATA.months.at(-2).id, comp: DATA.months.at(-1).id, empresa:'', sede:'', metric:'total_haberes', compareMetrics:['total_haberes','sueldo_liquido'], expandedCompanies:{}, expandedKpiRot:{}, expandedKpiAus:{}, expandedKpiLic:{}, kpiFocus:'', expandedDiffConcepts:{}, expandedDiffCompanies:{}, expandedDiffSedes:{}, conceptLimit:'15', conceptSort:'abs', conceptSearch:'', selectedConcept:'', detailFilters:{}};
let byId = Object.fromEntries(DATA.months.map(m=>[m.id,m]));
let detailFilterTimer = null;
let detailFilterKey = '';
const moneyMetrics = new Set(['total_haberes','sueldo_liquido','sueldo_base','hhee','gratificacion','movilizacion','colacion','total_descuentos','sb_promedio','liq_promedio']);
const M=n=>isNaN(+n)?'--':'$'+(+(+n/1e6).toFixed(1)).toLocaleString('es-CL')+'M';
const K=n=>isNaN(+n)?'--':'$'+Math.round(+n).toLocaleString('es-CL');
const N=n=>isNaN(+n)?'--':(+n).toLocaleString('es-CL');
const P=n=>isNaN(+n)?'--':(+n).toFixed(1)+'%';
const fmt=(key,n)=>(moneyMetrics.has(key)||isConceptMetric(key))?(Math.abs(n)>=1000000?M(n):K(n)):N(n);
const deltaPct=(a,b)=>a===0?(b===0?0:100):((b-a)/a*100);
const cls=n=>n>0?'dp':n<0?'dn':'d0';
const sign=n=>n>0?'+':'';
const esc=s=>String(s||'').replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/'/g,'\\&#39;');
const txt=s=>String(s??'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#39;');
const rotBdg=p=>p===0?`<span class="bdg bgr">0%</span>`:p<=5?`<span class="bdg bg">${p}%</span>`:p<=10?`<span class="bdg by">${p}%</span>`:`<span class="bdg br2">${p}%</span>`;
function label(id){return byId[id]?.label || id}
function metricOptions(){
  const core=[
    ['total_haberes','Suma Haberes'],['dotacion','Dotación'],['sueldo_liquido','Sueldo Líquido'],['sueldo_base','Sueldo Base'],
    ['gratificacion','Gratificación'],['movilizacion','Movilización'],['colacion','Colación'],['hhee','Horas Extras'],['total_descuentos','Total Rebajas']
  ];
  const conceptOptions=DATA.concept_options.map(c=>[`concept::${c}`,`Concepto: ${c}`]);
  return [...core,...conceptOptions];
}
function metricName(key){return Object.fromEntries(metricOptions())[key]||key}
function isConceptMetric(key=state.metric){return key.startsWith('concept::')}
function conceptName(key=state.metric){return key.replace('concept::','')}
function filterLabel(){return [state.empresa||'Todas las empresas',state.sede||'Todas las sedes'].join(' · ')}
function selectedRows(month=state.comp){
  return (DATA.groups.empresa_sede[month] || []).filter(r=>(!state.empresa||r.empresa===state.empresa)&&(!state.sede||r.sede===state.sede));
}
function getFilteredKpis(month){
  if(!state.empresa && !state.sede) return DATA.kpis[month];
  const rows=selectedRows(month);
  const sum=k=>rows.reduce((a,r)=>a+(r[k]||0),0),dot=sum('dotacion'),fte=sum('fte'),cero=sum('cero_dias'),menos=sum('menos30');
  return {
    dotacion:dot,fte,dias_prom:dot?+(rows.reduce((a,r)=>a+(r.dias_prom||0)*(r.dotacion||0),0)/dot).toFixed(1):0,
    cero_dias:cero,menos30:menos,rotacion_pct:dot?+((cero+menos)/dot*100).toFixed(1):0,ausentes_pct:dot?+(cero/dot*100).toFixed(1):0,
    total_haberes:sum('total_haberes'),sueldo_base:sum('sueldo_base'),sueldo_liquido:sum('sueldo_liquido'),total_descuentos:sum('total_descuentos'),
    gratificacion:sum('gratificacion'),hhee:sum('hhee'),licencia_dias:sum('licencia_dias'),ausentismo_dias:sum('ausentismo_dias'),licencia_personas:sum('licencia_personas'),ausentismo_personas:sum('ausentismo_personas'),movilizacion:sum('movilizacion'),colacion:sum('colacion'),
    sb_promedio:dot?Math.round(sum('sueldo_base')/dot):0,liq_promedio:dot?Math.round(sum('sueldo_liquido')/dot):0,pct_fte:dot?+(fte/dot*100).toFixed(1):0
  };
}
function movementStats(){
  const filter=r=>(!state.empresa||r.empresa===state.empresa)&&(!state.sede||r.sede===state.sede);
  const base=(DATA.details[state.base]||[]).filter(filter),comp=(DATA.details[state.comp]||[]).filter(filter);
  const bs=new Set(base.map(r=>r.rut)),cs=new Set(comp.map(r=>r.rut));
  return {ingresos:[...cs].filter(r=>!bs.has(r)).length,egresos:[...bs].filter(r=>!cs.has(r)).length,permanentes:[...cs].filter(r=>bs.has(r)).length};
}
function getConceptDict(month){
  const cg=DATA.concept_groups[month];
  if(state.empresa&&state.sede) return cg.empresa_sede[`${state.empresa} · ${state.sede}`]||{};
  if(state.empresa) return cg.empresa[state.empresa]||{};
  if(state.sede) return cg.sede[state.sede]||{};
  return cg.total||{};
}
function groupConceptValue(month,row,metric=state.metric){
  if(!isConceptMetric(metric)) return row[metric]||0;
  const key=`${row.empresa} · ${row.sede}`;
  return (DATA.concept_groups[month]?.empresa_sede?.[key]?.[conceptName(metric)]||0);
}
function metricValueForKpi(month,key=state.metric){
  if(isConceptMetric(key)) return getConceptDict(month)[conceptName(key)]||0;
  return getFilteredKpis(month)[key]||0;
}
function setOptions(sel, values, current){sel.innerHTML=values.map(v=>`<option value="${v.id}">${v.label}</option>`).join(''); sel.value=current}
function initSelectors(){
  const months=DATA.months;
  setOptions(selBase, months, state.base); setOptions(selComp, months, state.comp); setOptions(detMonth, months, state.comp);
  fillGlobalFilters();
  fillMetricOptions();
  selBase.onchange=()=>{state.base=selBase.value; renderAll()};
  selComp.onchange=()=>{state.comp=selComp.value; fillGlobalFilters(); syncDetailFilters(); renderAll()};
  selEmpresa.onchange=()=>{state.empresa=selEmpresa.value; state.sede=''; fillGlobalFilters(); syncDetailFilters(); renderAll()};
  selSede.onchange=()=>{state.sede=selSede.value; syncDetailFilters(); renderAll()};
  conceptLimit.onchange=()=>{state.conceptLimit=conceptLimit.value; renderConcepts(); renderDeviaciones()};
  conceptSort.onchange=()=>{state.conceptSort=conceptSort.value; renderConcepts()};
  conceptSearch.oninput=()=>{state.conceptSearch=conceptSearch.value.toLowerCase(); renderConcepts()};
  refreshDetailFilters();
}
function fillGlobalFilters(){
  const rows=DATA.groups.empresa_sede[state.comp]||[];
  const empresas=[...new Set(rows.map(r=>r.empresa))].sort();
  selEmpresa.innerHTML='<option value="">Todas las empresas</option>'+empresas.map(e=>`<option value="${e}">${e}</option>`).join('');
  if(empresas.includes(state.empresa)) selEmpresa.value=state.empresa; else state.empresa='';
  const sedes=[...new Set(rows.filter(r=>!state.empresa||r.empresa===state.empresa).map(r=>r.sede))].sort();
  selSede.innerHTML='<option value="">Todas las sedes</option>'+sedes.map(s=>`<option value="${s}">${s}</option>`).join('');
  if(sedes.includes(state.sede)) selSede.value=state.sede; else state.sede='';
}
function fillMetricOptions(){
  const opts=metricOptions();
  cmpMetricAdd.innerHTML=opts.map(([v,l])=>`<option value="${v}">${l}</option>`).join('');
}
function syncDetailFilters(){
  detMonth.value=state.comp;
  refreshDetailFilters();
  detEmpresa.value=state.empresa;
  refreshDetailFilters();
  detSede.value=state.sede;
}
function renderChips(){
  chipBase.textContent=label(state.base); chipComp.textContent=label(state.comp);
  sourceBadge.textContent=`Archivo cargado por defecto: ${DATA.metadata.source} · ${DATA.metadata.month_count} meses`;
  dashSub.textContent=`${DATA.metadata.record_count.toLocaleString('es-CL')} registros · ${DATA.metadata.company_count} empresas · ${DATA.metadata.month_count} meses · ${filterLabel()}`;
}
function renderKPIs(){
  const f=getFilteredKpis(state.base),m=getFilteredKpis(state.comp);
  const mov=movementStats();
  const cards=[
    ['Dotación','dotacion',m.dotacion,f.dotacion,m.dotacion-f.dotacion,'var(--blue)'],
    ['FTE','fte',m.fte,f.fte,m.fte-f.fte,'var(--green)'],
    ['Ingresos Mes','ingresos',mov.ingresos,0,mov.ingresos,'var(--green)'],
    ['Egresos Mes','egresos',mov.egresos,0,mov.egresos,'var(--red)'],
    ['Suma Haberes','total_haberes',m.total_haberes,f.total_haberes,m.total_haberes-f.total_haberes,'var(--purple)'],
    ['Sueldo Líquido','sueldo_liquido',m.sueldo_liquido,f.sueldo_liquido,m.sueldo_liquido-f.sueldo_liquido,'var(--green)'],
    ['Sueldo Base','sueldo_base',m.sueldo_base,f.sueldo_base,m.sueldo_base-f.sueldo_base,'var(--blue)'],
    ['Descuentos','total_descuentos',m.total_descuentos,f.total_descuentos,m.total_descuentos-f.total_descuentos,'var(--red)'],
    ['Gratificación','gratificacion',m.gratificacion,f.gratificacion,m.gratificacion-f.gratificacion,'var(--amber)'],
    ['Movilización','movilizacion',m.movilizacion,f.movilizacion,m.movilizacion-f.movilizacion,'var(--purple)'],
    ['Sueldo Base Prom.','sb_promedio',m.sb_promedio,f.sb_promedio,m.sb_promedio-f.sb_promedio,'var(--blue)'],
    ['Líquido Prom.','liq_promedio',m.liq_promedio,f.liq_promedio,m.liq_promedio-f.liq_promedio,'var(--green)'],
    ['Días Prom.','dias_prom',m.dias_prom,f.dias_prom,m.dias_prom-f.dias_prom,'var(--amber)'],
    ['Parciales','menos30',m.menos30,f.menos30,m.menos30-f.menos30,'var(--amber)'],
    ['% FTE','pct_fte',m.pct_fte,f.pct_fte,m.pct_fte-f.pct_fte,'var(--green)'],
  ];
  kpiGrid.innerHTML=cards.map(c=>{
    const [t,k,v,prev,d,color]=c;
    const isMovement=k==='ingresos'||k==='egresos';
    const delta=k.includes('pct')?`${sign(d)}${d.toFixed(1)}pp`:(k==='ingresos'||k==='egresos')?`${N(v)} personas`:`${sign(d)}${fmt(k,d)}`;
    return `<div class="kpi" style="border-top-color:${color}"><div class="kpi-lbl">${t}</div><div class="kpi-val" style="color:${color}">${k.includes('pct')?P(v):fmt(k,v)}</div><div class="kpi-foot"><span class="kpi-prev">${isMovement?`${label(state.base)} → ${label(state.comp)}`:`${label(state.base)}: ${k.includes('pct')?P(prev):fmt(k,prev)}`}</span><span class="kpi-dif ${isMovement?'':cls(d)}" style="${isMovement?`color:${color}`:''}">${delta}</span></div></div>`;
  }).join('');
}
function aggregateCompanyRows(month){
  const rows=selectedRows(month),by={};
  rows.forEach(r=>{
    const x=by[r.empresa]||(by[r.empresa]={empresa:r.empresa,key:r.empresa,dotacion:0,fte:0,cero_dias:0,menos30:0,total_haberes:0,sueldo_liquido:0,sueldo_base:0,hhee:0});
    ['dotacion','fte','cero_dias','menos30','total_haberes','sueldo_liquido','sueldo_base','hhee'].forEach(k=>x[k]+=r[k]||0);
  });
  return Object.values(by).map(r=>({...r,rotacion_pct:r.dotacion?+((r.cero_dias+r.menos30)/r.dotacion*100).toFixed(1):0})).sort((a,b)=>b.total_haberes-a.total_haberes);
}
function renderCompanyDivision(){
  const baseMap=Object.fromEntries(aggregateCompanyRows(state.base).map(r=>[r.empresa,r]));
  const compRows=aggregateCompanyRows(state.comp).sort((a,b)=>Math.abs((b.total_haberes||0)-(baseMap[b.empresa]?.total_haberes||0))-Math.abs((a.total_haberes||0)-(baseMap[a.empresa]?.total_haberes||0)));
  const sedeBase=Object.fromEntries(selectedRows(state.base).map(r=>[r.key,r]));
  const sedeComp=selectedRows(state.comp);
  const sedesByEmpresa=sedeComp.reduce((m,r)=>{(m[r.empresa]=m[r.empresa]||[]).push(r);return m;},{});
  companyDivisionBody.innerHTML=compRows.map(emp=>{
    const b=baseMap[emp.empresa]||{},d=(emp.total_haberes||0)-(b.total_haberes||0),open=!!state.expandedCompanies[emp.empresa];
    const empresaRow=`<tr class="clickable"><td><button class="exp-btn" onclick="event.stopPropagation();toggleCompany('${esc(emp.empresa)}')">${open?'−':'+'}</button></td><td class="tn" onclick="goDetail('${state.comp}','${esc(emp.empresa)}','')">${emp.empresa}</td><td class="nr">${N(emp.dotacion)}</td><td class="nr">${(+emp.fte||0).toFixed(2)}</td><td class="nr">${M(emp.total_haberes)}</td><td class="nr">${M(emp.sueldo_liquido)}</td><td class="nr">${M(emp.sueldo_base)}</td><td class="nr ${cls(d)}" style="font-weight:700">${sign(d)}${M(d)}</td></tr>`;
    const sedeRows=open?(sedesByEmpresa[emp.empresa]||[]).sort((a,b)=>Math.abs((b.total_haberes||0)-(sedeBase[b.key]?.total_haberes||0))-Math.abs((a.total_haberes||0)-(sedeBase[a.key]?.total_haberes||0))).map(s=>{
      const sb=sedeBase[s.key]||{},sd=(s.total_haberes||0)-(sb.total_haberes||0);
      return `<tr class="sede-row clickable" onclick="goDetail('${state.comp}','${esc(s.empresa)}','${esc(s.sede)}')"><td></td><td class="sede-name">${s.sede}</td><td class="nr">${N(s.dotacion)}</td><td class="nr">${(+s.fte||0).toFixed(2)}</td><td class="nr">${M(s.total_haberes)}</td><td class="nr">${M(s.sueldo_liquido)}</td><td class="nr">${M(s.sueldo_base)}</td><td class="nr ${cls(sd)}" style="font-weight:700">${sign(sd)}${M(sd)}</td></tr>`;
    }).join(''):'';
    return empresaRow+sedeRows;
  }).join('');
}
function toggleCompany(empresa){
  state.expandedCompanies[empresa]=!state.expandedCompanies[empresa];
  renderCompanyDivision();
}
function renderTotalsHighlights(){
  const f=getFilteredKpis(state.base),m=getFilteredKpis(state.comp);
  totalsSub.textContent=`${label(state.base)} → ${label(state.comp)} · ${filterLabel()}`;
  const rows=[
    ['Suma Haberes','total_haberes'],
    ['Sueldo Base','sueldo_base'],
    ['Gratificación','gratificacion'],
    ['Sueldo Líquido','sueldo_liquido'],
    ['Total Rebajas','total_descuentos'],
    ['Horas Extras','hhee'],
    ['Movilización','movilizacion'],
    ['Colación','colacion'],
    ['Dotación','dotacion'],
    ['FTE','fte'],
    ['Días parciales','menos30'],
  ];
  totalsBody.innerHTML=rows.sort((a,b)=>Math.abs((m[b[1]]||0)-(f[b[1]]||0))-Math.abs((m[a[1]]||0)-(f[a[1]]||0))).map(([labelTxt,key])=>{
    const d=(m[key]||0)-(f[key]||0),pp=deltaPct(f[key]||0,m[key]||0);
    return `<tr><td class="tn">${labelTxt}</td><td class="nr">${fmt(key,f[key]||0)}</td><td class="nr">${fmt(key,m[key]||0)}</td><td class="nr ${cls(d)}" style="font-weight:700">${sign(d)}${fmt(key,d)}</td><td class="nr ${cls(pp)}">${sign(pp)}${pp.toFixed(1)}%</td></tr>`;
  }).join('');
  const groupDiffs=compareGroupRows().map(r=>{
    const bv=r.base.total_haberes||0,cv=r.comp.total_haberes||0;
    return {...r,dif:cv-bv,pp:deltaPct(bv,cv),ddot:(r.comp.dotacion||0)-(r.base.dotacion||0)};
  });
  const topUp=[...groupDiffs].sort((a,b)=>b.dif-a.dif)[0];
  const topDown=[...groupDiffs].sort((a,b)=>a.dif-b.dif)[0];
  const conceptRows=getConceptRows('abs','',0);
  const topConcept=conceptRows[0];
  const habDif=m.total_haberes-f.total_haberes,liqDif=m.sueldo_liquido-f.sueldo_liquido,dotDif=m.dotacion-f.dotacion;
  highlights.innerHTML=[
    {t:'Variación neta de haberes',s:`${label(state.base)} → ${label(state.comp)}`,v:`${sign(habDif)}${M(habDif)}`,c:habDif},
    {t:'Variación de líquido',s:`Impacto final en pagos`,v:`${sign(liqDif)}${M(liqDif)}`,c:liqDif},
    {t:'Cambio de dotación',s:`Trabajadores del período`,v:`${sign(dotDif)}${N(dotDif)}`,c:dotDif},
    {t:'Mayor alza Empresa/Sede',s:topUp?.key||'Sin datos',v:topUp?`${sign(topUp.dif)}${M(topUp.dif)}`:'--',c:topUp?.dif||0},
    {t:'Mayor baja Empresa/Sede',s:topDown?.key||'Sin datos',v:topDown?`${sign(topDown.dif)}${M(topDown.dif)}`:'--',c:topDown?.dif||0},
    {t:'Concepto más explicativo',s:topConcept?.concepto||'Sin datos',v:topConcept?`${sign(topConcept.dif)}${M(topConcept.dif)}`:'--',c:topConcept?.dif||0},
  ].map(h=>`<div class="hi"><div class="hi-main"><div class="hi-title">${h.t}</div><div class="hi-sub">${h.s}</div></div><div class="hi-val ${cls(h.c)}">${h.v}</div></div>`).join('');
}
function renderTrend(){
  const cv=trend,ctx=cv.getContext('2d'),w=cv.width,h=cv.height,p=34;
  ctx.clearRect(0,0,w,h); ctx.font='11px Inter, sans-serif'; ctx.strokeStyle='#e2e6ef'; ctx.lineWidth=1;
  for(let i=0;i<4;i++){const y=p+i*(h-2*p)/3; ctx.beginPath(); ctx.moveTo(p,y); ctx.lineTo(w-p,y); ctx.stroke();}
  const vals=DATA.months.map(m=>getFilteredKpis(m.id).total_haberes),dots=DATA.months.map(m=>getFilteredKpis(m.id).dotacion);
  const maxV=Math.max(...vals),minV=Math.min(...vals),maxD=Math.max(...dots),minD=Math.min(...dots);
  const x=i=>p+i*(w-2*p)/(DATA.months.length-1);
  const yV=v=>h-p-(v-minV)/Math.max(maxV-minV,1)*(h-2*p);
  const yD=v=>h-p-(v-minD)/Math.max(maxD-minD,1)*(h-2*p);
  function line(arr,yf,color){ctx.beginPath();arr.forEach((v,i)=>i?ctx.lineTo(x(i),yf(v)):ctx.moveTo(x(i),yf(v)));ctx.strokeStyle=color;ctx.lineWidth=3;ctx.stroke();arr.forEach((v,i)=>{ctx.beginPath();ctx.arc(x(i),yf(v),3,0,Math.PI*2);ctx.fillStyle=color;ctx.fill();});}
  line(vals,yV,'#2563eb'); line(dots,yD,'#059669');
  DATA.months.forEach((m,i)=>{ctx.fillStyle='#9198b5';ctx.textAlign='center';ctx.fillText(m.label.split(' ')[0],x(i),h-10);});
  ctx.textAlign='left';ctx.fillStyle='#2563eb';ctx.fillText('Suma Haberes',p,15);ctx.fillStyle='#059669';ctx.fillText('Dotación',p+98,15);
}
function compareGroupRows(){
  const base=selectedRows(state.base),comp=selectedRows(state.comp);
  const bm=Object.fromEntries(base.map(r=>[r.key,r])),cm=Object.fromEntries(comp.map(r=>[r.key,r]));
  const sortMetric=state.compareMetrics?.[0]||state.metric;
  return [...new Set([...base.map(r=>r.key),...comp.map(r=>r.key)])].map(k=>({key:k,base:bm[k]||{},comp:cm[k]||{}})).sort((a,b)=>{
    const da=groupConceptValue(state.comp,a.comp,sortMetric)-groupConceptValue(state.base,a.base,sortMetric);
    const db=groupConceptValue(state.comp,b.comp,sortMetric)-groupConceptValue(state.base,b.base,sortMetric);
    return Math.abs(db)-Math.abs(da);
  });
}
function renderComparativo(){
  renderCompareMetricChips();
  const primary=state.compareMetrics[0]||'total_haberes';
  const dif=metricValueForKpi(state.comp,primary)-metricValueForKpi(state.base,primary);
  cmpBanner.innerHTML=`<div><div class="cbl">Diferencia ${metricName(primary)}</div><div class="cbv">${sign(dif)}${fmt(primary,dif)}</div><div class="cbs">${label(state.base)} → ${label(state.comp)} · ${sign(deltaPct(metricValueForKpi(state.base,primary),metricValueForKpi(state.comp,primary)))}${deltaPct(metricValueForKpi(state.base,primary),metricValueForKpi(state.comp,primary)).toFixed(1)}%</div></div><div><div class="cbl">Columnas comparadas</div><div class="cbv">${state.compareMetrics.length}</div><div class="cbs">${filterLabel()}</div></div>`;
  const first='<th>Empresa</th><th>Sede</th>';
  const metricHeads=state.compareMetrics.map(k=>`<th class="nr">${metricName(k)} ${label(state.base)}</th><th class="nr">${metricName(k)} ${label(state.comp)}</th><th class="nr">Δ ${metricName(k)}</th>`).join('');
  cmpHead.innerHTML=`${first}<th class="nr">Dot. Base</th><th class="nr">Dot. Comp.</th>${metricHeads}`;
  const totals=Object.fromEntries(state.compareMetrics.map(k=>[k,{b:0,c:0}]));
  let db=0,dc=0;
  cmpBody.innerHTML=compareGroupRows().map(({key,base,comp})=>{
    db+=(base.dotacion||0);dc+=(comp.dotacion||0);
    const emp=comp.empresa||base.empresa||'',sed=comp.sede||base.sede||'';
    const name=`<td class="tn">${emp}</td><td>${sed}</td>`;
    const cells=state.compareMetrics.map(k=>{
      const bv=groupConceptValue(state.base,base,k),cv=groupConceptValue(state.comp,comp,k),d=cv-bv;
      totals[k].b+=bv;totals[k].c+=cv;
      return `<td class="nr">${fmt(k,bv)}</td><td class="nr">${fmt(k,cv)}</td><td class="nr ${cls(d)}" style="font-weight:700">${sign(d)}${fmt(k,d)}</td>`;
    }).join('');
    return `<tr class="clickable" onclick="goDetail('${state.comp}','${esc(emp)}','${esc(sed)}')">${name}<td class="nr">${N(base.dotacion||0)}</td><td class="nr">${N(comp.dotacion||0)}</td>${cells}</tr>`;
  }).join('');
  const totalCells=state.compareMetrics.map(k=>{const t=totals[k],d=t.c-t.b;return `<td class="nr"><b>${fmt(k,t.b)}</b></td><td class="nr"><b>${fmt(k,t.c)}</b></td><td class="nr ${cls(d)}"><b>${sign(d)}${fmt(k,d)}</b></td>`}).join('');
  cmpFoot.innerHTML=`<tr><td colspan="2" style="padding:9px 11px"><b>Total</b></td><td class="nr"><b>${N(db)}</b></td><td class="nr"><b>${N(dc)}</b></td>${totalCells}</tr>`;
}
function renderCompareMetricChips(){
  cmpMetricChips.innerHTML=state.compareMetrics.map(k=>`<span class="chip-x">${metricName(k)}${state.compareMetrics.length>1?`<button onclick="removeCompareMetric('${esc(k)}')">×</button>`:''}</span>`).join('');
}
function addCompareMetric(){
  const key=cmpMetricAdd.value;
  if(key&&!state.compareMetrics.includes(key)) state.compareMetrics.push(key);
  renderComparativo();
}
function removeCompareMetric(key){
  state.compareMetrics=state.compareMetrics.filter(k=>k!==key);
  renderComparativo();
}
function conceptType(concept){return DATA.concept_types?.[concept]||'haber';}
const requiredMonthlyDiffConcepts=['Horas Extras Empresa 50%'];
function getConceptRows(sort=state.conceptSort, search=state.conceptSearch, limitValue=state.conceptLimit, type=''){
  const a=getConceptDict(state.base),b=getConceptDict(state.comp);
  let rows=[...new Set([...Object.keys(a),...Object.keys(b)])].map(c=>({concepto:c,base:a[c]||0,comp:b[c]||0,dif:(b[c]||0)-(a[c]||0)})).filter(r=>r.base||r.comp);
  if(type) rows=rows.filter(r=>conceptType(r.concepto)===type);
  if(search) rows=rows.filter(r=>r.concepto.toLowerCase().includes(search));
  rows.sort((x,y)=>{
    if(sort==='up') return y.dif-x.dif;
    if(sort==='down') return x.dif-y.dif;
    if(sort==='name') return x.concepto.localeCompare(y.concepto);
    return Math.abs(y.dif)-Math.abs(x.dif);
  });
  const fullRows=rows;
  if(limitValue && limitValue!=='all') rows=rows.slice(0,+limitValue);
  if(!search && (!type||type==='haber')){
    requiredMonthlyDiffConcepts.forEach(concept=>{
      const required=fullRows.find(r=>r.concepto===concept);
      if(required&&!rows.some(r=>r.concepto===concept)) rows.push(required);
    });
  }
  return rows;
}
function renderConcepts(){
  ccBase.textContent=label(state.base); ccComp.textContent=label(state.comp);
  const rows=getConceptRows(state.conceptSort,state.conceptSearch,state.conceptLimit,'haber');
  const max=Math.max(...rows.map(r=>Math.abs(r.dif)),1),totAbs=rows.reduce((s,r)=>s+Math.abs(r.dif),0),sumHab=rows.reduce((s,r)=>s+r.dif,0);
  conceptSub.textContent=`${rows.length} haberes visibles · expande concepto / empresa / sede / persona · ${label(state.base)} → ${label(state.comp)} · ${filterLabel()}`;
  ccRows.innerHTML=rows.length?rows.map(r=>renderConceptTreeRow(r,max,totAbs)).join(''):`<div class="cc-section">Haberes<span>Sin diferencias visibles</span></div>`;
  const totalDif=getFilteredKpis(state.comp).total_haberes-getFilteredKpis(state.base).total_haberes;
  const labelLimit=state.conceptLimit==='all'?'todos los conceptos':`top ${state.conceptLimit}`;
  ccTot.innerHTML=`<div style="display:flex;align-items:center;justify-content:space-between;font-size:12px;flex-wrap:wrap;gap:8px"><b>Suma haberes ${labelLimit} visibles</b><div style="display:flex;gap:20px;align-items:center;flex-wrap:wrap"><span class="bdg bb">Conceptos visibles: ${sign(sumHab)}${M(sumHab)}</span><span class="bdg bb">Δ total haberes: ${sign(totalDif)}${M(totalDif)}</span></div></div>`;
}
function renderConceptTreeRow(r,max,totAbs){
  const color=r.dif>=0?'var(--green)':'var(--red)',pct=totAbs?Math.abs(r.dif)/totAbs*100:0,conceptKey=esc(r.concepto),open=!!state.expandedDiffConcepts[r.concepto];
  const row=`<div class="cc-row clickable ${open?'selected-line':''}"><div class="cc-name"><button class="exp-btn" onclick="event.stopPropagation();toggleDiffConcept('${conceptKey}')">${open?'−':'+'}</button><span style="margin-left:8px">${r.concepto}</span></div><div class="cc-num">${M(r.base)}</div><div class="cc-num">${M(r.comp)}</div><div class="cc-num" style="color:${color};font-weight:700">${sign(r.dif)}${M(r.dif)}</div><div class="cc-bar"><div class="cc-mb"><div class="cc-mf" style="width:${(Math.abs(r.dif)/max*100).toFixed(1)}%;background:${color}"></div></div><span style="font-size:10.5px;color:var(--text3);font-family:var(--m);min-width:34px">${pct.toFixed(1)}%</span></div></div>`;
  if(!open) return row;
  const tree=buildConceptTree(workerConceptRows(state.base,r.concepto),workerConceptRows(state.comp,r.concepto));
  return row+tree.map(emp=>renderConceptCompanyRow(r.concepto,emp)).join('');
}
function renderConceptCompanyRow(concept,emp){
  const key=`${concept}||${emp.empresa}`,safe=esc(key),open=!!state.expandedDiffCompanies[key];
  const row=`<div class="cc-row clickable"><div class="cc-name" style="padding-left:26px"><button class="exp-btn" onclick="event.stopPropagation();toggleDiffCompany('${safe}')">${open?'−':'+'}</button><span class="tn" style="margin-left:8px" onclick="goDetail('${state.comp}','${esc(emp.empresa)}','')">${emp.empresa}</span></div><div class="cc-num">${M(emp.base)}</div><div class="cc-num">${M(emp.comp)}</div><div class="cc-num ${cls(emp.dif)}" style="font-weight:700">${sign(emp.dif)}${M(emp.dif)}</div><div class="cc-num ${cls(emp.pp)}">${sign(emp.pp)}${emp.pp.toFixed(1)}%</div></div>`;
  return row+(open?emp.sedes.map(s=>renderConceptSedeRow(concept,emp.empresa,s)).join(''):'');
}
function renderConceptSedeRow(concept,empresa,s){
  const key=`${concept}||${empresa}||${s.sede}`,safe=esc(key),open=!!state.expandedDiffSedes[key];
  const row=`<div class="cc-row clickable"><div class="cc-name" style="padding-left:52px"><button class="exp-btn" onclick="event.stopPropagation();toggleDiffSede('${safe}')">${open?'−':'+'}</button><span class="sede-name" style="padding-left:8px" onclick="goDetail('${state.comp}','${esc(empresa)}','${esc(s.sede)}')">${s.sede}</span></div><div class="cc-num">${M(s.base)}</div><div class="cc-num">${M(s.comp)}</div><div class="cc-num ${cls(s.dif)}" style="font-weight:700">${sign(s.dif)}${M(s.dif)}</div><div class="cc-num ${cls(s.pp)}">${sign(s.pp)}${s.pp.toFixed(1)}%</div></div>`;
  return row+(open?s.workers.map(w=>`<div class="cc-row clickable" onclick="goWorkerDetail('${state.comp}','${esc(w.rut)}')"><div class="cc-name" style="padding-left:92px"><span style="font-family:var(--m);font-size:11px;color:var(--text3)">${txt(w.rut)}</span> · ${txt(w.nombre)}</div><div class="cc-num">${M(w.base)}</div><div class="cc-num">${M(w.comp)}</div><div class="cc-num ${cls(w.dif)}" style="font-weight:700">${sign(w.dif)}${M(w.dif)}</div><div class="cc-num ${cls(w.pp)}">${sign(w.pp)}${w.pp.toFixed(1)}%</div></div>`).join(''):'');
}
function toggleDiffConcept(concept){
  state.expandedDiffConcepts[concept]=!state.expandedDiffConcepts[concept];
  renderConcepts();
}
function workerConceptRows(month,concept){
  return (DATA.details[month]||[]).filter(r=>(!state.empresa||r.empresa===state.empresa)&&(!state.sede||r.sede===state.sede)).map(r=>({
    rut:r.rut,nombre:r.nombre,empresa:r.empresa,sede:r.sede,value:+(r.concepts?.[concept]||0)
  })).filter(r=>r.value!==0);
}
function buildConceptTree(baseRows,compRows){
  const byRut={};
  baseRows.forEach(r=>{byRut[r.rut]={...r,base:r.value,comp:0}});
  compRows.forEach(r=>{byRut[r.rut]={...(byRut[r.rut]||r),...r,base:byRut[r.rut]?.base||0,comp:r.value}});
  const empMap={};
  Object.values(byRut).forEach(w=>{
    const emp=empMap[w.empresa]||(empMap[w.empresa]={empresa:w.empresa,base:0,comp:0,sedeMap:{}});
    const sede=emp.sedeMap[w.sede]||(emp.sedeMap[w.sede]={sede:w.sede,base:0,comp:0,workers:[]});
    const item={...w,dif:w.comp-w.base,pp:deltaPct(w.base,w.comp)};
    emp.base+=w.base;emp.comp+=w.comp;sede.base+=w.base;sede.comp+=w.comp;sede.workers.push(item);
  });
  return Object.values(empMap).map(emp=>{
    const sedes=Object.values(emp.sedeMap).map(s=>({...s,dif:s.comp-s.base,pp:deltaPct(s.base,s.comp),workers:s.workers.sort((a,b)=>Math.abs(b.dif)-Math.abs(a.dif))})).sort((a,b)=>Math.abs(b.dif)-Math.abs(a.dif));
    return {...emp,dif:emp.comp-emp.base,pp:deltaPct(emp.base,emp.comp),sedes};
  }).sort((a,b)=>Math.abs(b.dif)-Math.abs(a.dif));
}
function toggleDiffCompany(empresa){
  state.expandedDiffCompanies[empresa]=!state.expandedDiffCompanies[empresa];
  renderConcepts();
}
function toggleDiffSede(sedeKey){
  state.expandedDiffSedes[sedeKey]=!state.expandedDiffSedes[sedeKey];
  renderConcepts();
}
function filteredDetails(month){
  return (DATA.details[month]||[]).filter(r=>(!state.empresa||r.empresa===state.empresa)&&(!state.sede||r.sede===state.sede));
}
function treeFromWorkers(workers,valueKey){
  const emps={};
  workers.forEach(w=>{
    const emp=emps[w.empresa]||(emps[w.empresa]={empresa:w.empresa,count:0,value:0,sedeMap:{}});
    const sede=emp.sedeMap[w.sede]||(emp.sedeMap[w.sede]={sede:w.sede,count:0,value:0,workers:[]});
    emp.count++;sede.count++;emp.value+=(+w[valueKey]||0);sede.value+=(+w[valueKey]||0);sede.workers.push(w);
  });
  return Object.values(emps).map(e=>{
    const sedes=Object.values(e.sedeMap).map(s=>({...s,workers:s.workers.sort((a,b)=>(+b[valueKey]||0)-(+a[valueKey]||0))})).sort((a,b)=>b.value-a.value||b.count-a.count);
    return {...e,sedes};
  }).sort((a,b)=>b.value-a.value||b.count-a.count);
}
function renderWorkerTree(body,tree,expanded,prefix,mode){
  body.innerHTML=tree.map(emp=>{
    const eKey=`${prefix}:emp:${emp.empresa}`,eOpen=!!expanded[eKey];
    const empRow=`<tr class="clickable"><td><button class="exp-btn" onclick="event.stopPropagation();toggleKpiNode('${esc(eKey)}','${prefix}')">${eOpen?'−':'+'}</button></td><td class="tn">${emp.empresa}</td>${mode==='rot'?`<td><span class="bdg bb">${N(emp.count)} mov.</span></td><td class="nr"></td><td class="nr"></td><td class="nr">${M(emp.value)}</td>`:`<td class="nr">${N(emp.value)}</td><td class="nr"></td><td class="nr"></td>`}</tr>`;
    const sedeRows=eOpen?emp.sedes.map(s=>{
      const sKey=`${prefix}:sede:${emp.empresa} · ${s.sede}`,sOpen=!!expanded[sKey];
      const sedeRow=`<tr class="sede-row clickable"><td><button class="exp-btn" onclick="event.stopPropagation();toggleKpiNode('${esc(sKey)}','${prefix}')">${sOpen?'−':'+'}</button></td><td class="sede-name">${s.sede}</td>${mode==='rot'?`<td><span class="bdg bb">${N(s.count)} mov.</span></td><td class="nr"></td><td class="nr"></td><td class="nr">${M(s.value)}</td>`:`<td class="nr">${N(s.value)}</td><td class="nr"></td><td class="nr"></td>`}</tr>`;
      const workerRows=sOpen?s.workers.map(w=>{
        const name=`<span style="font-family:var(--m);font-size:11px;color:var(--text3)">${txt(w.rut)}</span> · ${txt(w.nombre)}`;
        if(mode==='rot') return `<tr class="sede-row clickable" onclick="goWorkerDetail('${w.month}','${esc(w.rut)}')"><td></td><td class="sede-name" style="padding-left:58px">${name}</td><td><span class="bdg ${w.tipo==='Ingreso'?'bg':'br2'}">${w.tipo}</span></td><td class="nr">${w.dias}</td><td class="nr">${(+w.fte||0).toFixed(2)}</td><td class="nr">${K(w.total_haberes)}</td></tr>`;
        return `<tr class="sede-row clickable" onclick="goWorkerDetail('${state.comp}','${esc(w.rut)}')"><td></td><td class="sede-name" style="padding-left:58px">${name}</td><td class="nr">${N(w.value)}</td><td class="nr">${w.dias}</td><td class="nr">${(+w.fte||0).toFixed(2)}</td></tr>`;
      }).join(''):'';
      return sedeRow+workerRows;
    }).join(''):'';
    return empRow+sedeRows;
  }).join('');
}
function expandWorkerTree(tree,expanded,prefix){
  tree.forEach(emp=>{
    expanded[`${prefix}:emp:${emp.empresa}`]=true;
    emp.sedes.forEach(s=>expanded[`${prefix}:sede:${emp.empresa} · ${s.sede}`]=true);
  });
}
function selectKpiDetail(kind){
  state.kpiFocus=kind;
  renderKpisMenu();
  const target=kind==='aus'?kpiAusBody:kind==='lic'?kpiLicBody:kpiRotBody;
  target.closest('.panel')?.scrollIntoView({behavior:'smooth',block:'start'});
}
function toggleKpiNode(key,prefix){
  const map=prefix==='rot'?state.expandedKpiRot:prefix==='aus'?state.expandedKpiAus:state.expandedKpiLic;
  map[key]=!map[key];
  renderKpisMenu();
}
function renderKpisMenu(){
  const base=filteredDetails(state.base),comp=filteredDetails(state.comp),bm=Object.fromEntries(base.map(r=>[r.rut,r])),cm=Object.fromEntries(comp.map(r=>[r.rut,r]));
  const ingresos=comp.filter(r=>!bm[r.rut]).map(r=>({...r,month:state.comp,tipo:'Ingreso',value:r.total_haberes}));
  const egresos=base.filter(r=>!cm[r.rut]).map(r=>({...r,month:state.base,tipo:'Egreso',value:r.total_haberes}));
  const rot=[...ingresos,...egresos].sort((a,b)=>b.value-a.value);
  const aus=comp.filter(r=>(+r.ausentismo_dias||0)>0).map(r=>({...r,value:+r.ausentismo_dias||0}));
  const lic=comp.filter(r=>(+r.licencia_dias||0)>0).map(r=>({...r,value:+r.licencia_dias||0}));
  const f=getFilteredKpis(state.base),m=getFilteredKpis(state.comp);
  kpiOpsSub.textContent=`${label(state.base)} → ${label(state.comp)} · ${filterLabel()}`;
  const cards=[
    {id:'ingresos',t:'Ingresos',v:N(ingresos.length),s:'Click: detalle de personas nuevas en mes comparación',c:'var(--green)'},
    {id:'egresos',t:'Egresos',v:N(egresos.length),s:'Click: detalle de personas que estaban en mes base',c:'var(--red)'},
    {id:'aus',t:'Ausentismo',v:N(m.ausentismo_dias||0),s:`Click: ${N(m.ausentismo_personas||aus.length)} personas · Δ ${sign((m.ausentismo_dias||0)-(f.ausentismo_dias||0))}${N((m.ausentismo_dias||0)-(f.ausentismo_dias||0))} días`,c:'var(--amber)'},
    {id:'lic',t:'Licencias',v:N(m.licencia_dias||0),s:`Click: ${N(m.licencia_personas||lic.length)} personas · Δ ${sign((m.licencia_dias||0)-(f.licencia_dias||0))}${N((m.licencia_dias||0)-(f.licencia_dias||0))} días`,c:'var(--blue)'},
  ];
  kpiOpsGrid.innerHTML=cards.map(x=>`<div class="rc clickable ${state.kpiFocus===x.id?'active':''}" onclick="selectKpiDetail('${x.id}')" style="border-top:3px solid ${x.c}"><div class="rv" style="color:${x.c}">${x.v}</div><div class="rl">${x.t}</div><div class="rs">${x.s}</div></div>`).join('');
  const rotView=state.kpiFocus==='ingresos'?ingresos:state.kpiFocus==='egresos'?egresos:rot;
  const rotTree=treeFromWorkers(rotView,'value'),ausTree=treeFromWorkers(aus,'value'),licTree=treeFromWorkers(lic,'value');
  if(state.kpiFocus==='ingresos'||state.kpiFocus==='egresos'){state.expandedKpiRot={};expandWorkerTree(rotTree,state.expandedKpiRot,'rot');}
  if(state.kpiFocus==='aus'){state.expandedKpiAus={};expandWorkerTree(ausTree,state.expandedKpiAus,'aus');}
  if(state.kpiFocus==='lic'){state.expandedKpiLic={};expandWorkerTree(licTree,state.expandedKpiLic,'lic');}
  renderWorkerTree(kpiRotBody,rotTree,state.expandedKpiRot,'rot','rot');
  renderWorkerTree(kpiAusBody,ausTree,state.expandedKpiAus,'aus','aus');
  renderWorkerTree(kpiLicBody,licTree,state.expandedKpiLic,'lic','lic');
}
function renderDeviaciones(){
  const f=getFilteredKpis(state.base),m=getFilteredKpis(state.comp);
  const base=(DATA.details[state.base]||[]).filter(r=>(!state.empresa||r.empresa===state.empresa)&&(!state.sede||r.sede===state.sede));
  const comp=(DATA.details[state.comp]||[]).filter(r=>(!state.empresa||r.empresa===state.empresa)&&(!state.sede||r.sede===state.sede));
  const bs=new Set(base.map(r=>r.rut)),cs=new Set(comp.map(r=>r.rut));
  const bajas=[...bs].filter(r=>!cs.has(r)).length,altas=[...cs].filter(r=>!bs.has(r)).length,perm=[...bs].filter(r=>cs.has(r)).length;
  const habDif=m.total_haberes-f.total_haberes,pp=deltaPct(f.total_haberes,m.total_haberes);
  devGrid.innerHTML=`<div class="rc" style="border-top:3px solid ${habDif>=0?'var(--green)':'var(--red)'}"><div class="rv" style="color:${habDif>=0?'var(--green)':'var(--red)'}">${sign(habDif)}${M(habDif)}</div><div class="rl">Desviación neta haberes</div><div class="rs">${sign(pp)}${pp.toFixed(1)}% vs ${label(state.base)}</div></div><div class="rc" style="border-top:3px solid var(--blue)"><div class="rv" style="color:var(--blue)">${N(perm)}</div><div class="rl">RUTs permanentes</div><div class="rs">${N(altas)} altas · ${N(bajas)} bajas</div></div><div class="rc" style="border-top:3px solid var(--amber)"><div class="rv" style="color:var(--amber)">${sign(m.sb_promedio-f.sb_promedio)}${K(m.sb_promedio-f.sb_promedio)}</div><div class="rl">Cambio sueldo base prom.</div><div class="rs">${K(f.sb_promedio)} → ${K(m.sb_promedio)}</div></div>`;
  const diffs=compareGroupRows().map(r=>{
    const bv=r.base.total_haberes||0,cv=r.comp.total_haberes||0;
    return {...r,dif:cv-bv,pp:deltaPct(bv,cv),ddot:(r.comp.dotacion||0)-(r.base.dotacion||0)};
  });
  const rowHtml=r=>`<tr class="clickable" onclick="goDetail('${state.comp}','${esc(r.comp.empresa||r.base.empresa||'')}','${esc(r.comp.sede||r.base.sede||'')}')"><td class="tn">${r.key}</td><td class="nr">${M(r.base.total_haberes||0)}</td><td class="nr">${M(r.comp.total_haberes||0)}</td><td class="nr ${cls(r.dif)}" style="font-weight:700">${sign(r.dif)}${M(r.dif)}</td><td class="nr ${cls(r.pp)}">${sign(r.pp)}${r.pp.toFixed(1)}%</td><td class="nr ${cls(r.ddot)}">${sign(r.ddot)}${N(r.ddot)}</td></tr>`;
  devUpSub.textContent=`${label(state.base)} → ${label(state.comp)}`;
  devDownSub.textContent=`${label(state.base)} → ${label(state.comp)}`;
  devUpBody.innerHTML=[...diffs].filter(r=>r.dif>0).sort((a,b)=>b.dif-a.dif).slice(0,12).map(rowHtml).join('');
  devDownBody.innerHTML=[...diffs].filter(r=>r.dif<0).sort((a,b)=>a.dif-b.dif).slice(0,12).map(rowHtml).join('');
  const conceptRows=[...getConceptRows('abs','',0)].sort((a,b)=>Math.abs(b.dif)-Math.abs(a.dif)).slice(0,20);
  devConceptBody.innerHTML=conceptRows.map(r=>{const p=deltaPct(r.base,r.comp);return `<tr><td class="tn">${r.concepto}</td><td class="nr">${M(r.base)}</td><td class="nr">${M(r.comp)}</td><td class="nr ${cls(r.dif)}" style="font-weight:700">${sign(r.dif)}${M(r.dif)}</td><td class="nr ${cls(p)}">${sign(p)}${p.toFixed(1)}%</td></tr>`}).join('');
  const ratios=[
    ['Haberes por trabajador','total_haberes_dot',f.total_haberes/Math.max(f.dotacion,1),m.total_haberes/Math.max(m.dotacion,1),'$'],
    ['Líquido por trabajador','liquido_dot',f.sueldo_liquido/Math.max(f.dotacion,1),m.sueldo_liquido/Math.max(m.dotacion,1),'$'],
    ['Sueldo base promedio','sb_promedio',f.sb_promedio,m.sb_promedio,'$'],
    ['% FTE','pct_fte',f.pct_fte,m.pct_fte,'%'],
    ['% Ausentes','ausentes_pct',f.ausentes_pct,m.ausentes_pct,'%'],
    ['Días promedio','dias_prom',f.dias_prom,m.dias_prom,'n'],
  ];
  devRatioBody.innerHTML=ratios.map(([name,key,b,c,type])=>{const d=c-b;const baseFmt=type==='$'?K(b):type==='%'?P(b):b.toFixed(1);const compFmt=type==='$'?K(c):type==='%'?P(c):c.toFixed(1);const dFmt=type==='$'?`${sign(d)}${K(d)}`:type==='%'?`${sign(d)}${d.toFixed(1)}pp`:`${sign(d)}${d.toFixed(1)}`;return `<tr><td class="tn">${name}</td><td class="nr">${baseFmt}</td><td class="nr">${compFmt}</td><td class="nr ${cls(d)}" style="font-weight:700">${dFmt}</td></tr>`}).join('');
}
function auditRows(){
  return (DATA.details[state.comp]||[]).filter(r=>(!state.empresa||r.empresa===state.empresa)&&(!state.sede||r.sede===state.sede));
}
function isManipPae(r){
  return String(r.cargo||'').toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'').includes('manipulador (a) de alimentos pae');
}
function renderAuditoria(){
  const rows=auditRows(),manip=rows.filter(isManipPae);
  const zero=rows.filter(r=>+r.dias===0 && +r.total_haberes!==0);
  const expected=r=>Math.round(539000*Math.max(0,Math.min(+r.dias||0,30))/30);
  const baseOk=manip.filter(r=>Math.abs((+r.sueldo_base||0)-expected(r))<=1);
  const withBono=manip.filter(r=>(+r.bono_manip_pae||0)>0);
  auditSub.textContent=`${label(state.comp)} · ${filterLabel()}`;
  auditGrid.innerHTML=`<div class="rc" style="border-top:3px solid ${zero.length?'var(--red)':'var(--green)'}"><div class="rv" style="color:${zero.length?'var(--red)':'var(--green)'}">${N(zero.length)}</div><div class="rl">0 días con haberes</div><div class="rs">Total observado: ${M(zero.reduce((a,r)=>a+r.total_haberes,0))}</div></div><div class="rc" style="border-top:3px solid var(--blue)"><div class="rv" style="color:var(--blue)">${N(baseOk.length)} / ${N(manip.length)}</div><div class="rl">Base $539.000 proporcional OK</div><div class="rs">${P(manip.length?baseOk.length/manip.length*100:0)} de manipuladores PAE</div></div><div class="rc" style="border-top:3px solid var(--green)"><div class="rv" style="color:var(--green)">${N(withBono.length)}</div><div class="rl">Manipuladores PAE con bono</div><div class="rs">Monto: ${M(withBono.reduce((a,r)=>a+(+r.bono_manip_pae||0),0))}</div></div>`;
  auditZeroBody.innerHTML=zero.sort((a,b)=>b.total_haberes-a.total_haberes).map(r=>`<tr class="clickable" onclick="goDetail('${state.comp}','${esc(r.empresa)}','${esc(r.sede)}')"><td>${r.empresa}</td><td><span class="bdg bb">${r.sede}</span></td><td style="font-family:var(--m);font-size:11px;color:var(--text3)">${r.rut}</td><td class="tn">${r.nombre}</td><td>${r.cargo}</td><td class="nr">${r.dias}</td><td class="nr">${K(r.total_haberes)}</td><td class="nr">${K(r.sueldo_liquido)}</td></tr>`).join('');
  const bySede=manip.reduce((m,r)=>{const k=`${r.empresa} · ${r.sede}`;const x=m[k]||(m[k]={empresa:r.empresa,sede:r.sede,total:0,ok:0,diff:0,personas:0,conBono:0,montoBono:0});x.total++;if(Math.abs((+r.sueldo_base||0)-expected(r))<=1)x.ok++;else x.diff++;x.personas++;if((+r.bono_manip_pae||0)>0)x.conBono++;x.montoBono+=(+r.bono_manip_pae||0);return m;},{});
  const sedeRows=Object.values(bySede).sort((a,b)=>b.total-a.total);
  let tt=0,to=0,td=0,tp=0,tb=0,tmb=0;
  auditBaseBody.innerHTML=sedeRows.map(r=>{tt+=r.total;to+=r.ok;td+=r.diff;return `<tr class="clickable" onclick="goDetail('${state.comp}','${esc(r.empresa)}','${esc(r.sede)}')"><td class="tn">${r.empresa}</td><td><span class="bdg bb">${r.sede}</span></td><td class="nr">${N(r.total)}</td><td class="nr"><span class="bdg bg">${N(r.ok)}</span></td><td class="nr"><span class="bdg ${r.diff?'br2':'bgr'}">${N(r.diff)}</span></td><td class="nr">${P(r.total?r.ok/r.total*100:0)}</td></tr>`}).join('');
  auditBaseFoot.innerHTML=`<tr><td colspan="2" style="padding:9px 11px"><b>Total</b></td><td class="nr"><b>${N(tt)}</b></td><td class="nr"><b>${N(to)}</b></td><td class="nr"><b>${N(td)}</b></td><td class="nr"><b>${P(tt?to/tt*100:0)}</b></td></tr>`;
  auditBonoBody.innerHTML=sedeRows.map(r=>{tp+=r.personas;tb+=r.conBono;tmb+=r.montoBono;return `<tr class="clickable" onclick="goDetail('${state.comp}','${esc(r.empresa)}','${esc(r.sede)}')"><td class="tn">${r.empresa}</td><td><span class="bdg bb">${r.sede}</span></td><td class="nr">${N(r.personas)}</td><td class="nr">${N(r.conBono)}</td><td class="nr">${P(r.personas?r.conBono/r.personas*100:0)}</td><td class="nr">${M(r.montoBono)}</td></tr>`}).join('');
  auditBonoFoot.innerHTML=`<tr><td colspan="2" style="padding:9px 11px"><b>Total</b></td><td class="nr"><b>${N(tp)}</b></td><td class="nr"><b>${N(tb)}</b></td><td class="nr"><b>${P(tp?tb/tp*100:0)}</b></td><td class="nr"><b>${M(tmb)}</b></td></tr>`;
}
const downloadReports=[
  {id:'general',name:'General',file:'Revision_General',rule:'Todas las empresas y áreas',filter:r=>true},
  {id:'rrhh',name:'RRHH',file:'Revision_RRHH',rule:'Sin Pakarati y sin cargos Gerente General',filter:r=>!isPakaratiRow(r)&&!String(r.cargo||'').toLowerCase().includes('gerente general')},
  {id:'planta',name:'Planta',file:'Revision_PLANTA',rule:'Solo sede Planta Crux',filter:r=>String(r.sede||'').toLowerCase()==='planta crux'},
  {id:'pakarati',name:'Pakarati',file:'Revision_PAKARATI',rule:'Solo empresa Claudia Pakarati',filter:r=>isPakaratiRow(r)},
  {id:'jcuevas',name:'J.Cuevas',file:'Revision_JCUEVAS',rule:'Sedes UT, excluyendo Alianza y Claudia Pakarati',filter:r=>String(r.sede||'').toUpperCase().startsWith('UT')&&!isAlianzaRow(r)&&!isPakaratiRow(r)},
  {id:'alianza',name:'Alianza',file:'Revision_ALIANZA',rule:'Solo empresa Alianza',filter:r=>isAlianzaRow(r)},
];
function norm(s){return String(s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toLowerCase();}
function isPakarati(s){const x=norm(s);return x.includes('pakarati')||x.includes('claudia lorena');}
function isAlianza(s){return norm(s).includes('alianza');}
function companyText(r){return [r.empresa,r.raw?.[0],r.raw?.[1],r.raw?.[2]].map(x=>String(x||'')).join(' ');}
function isPakaratiRow(r){return isPakarati(companyText(r));}
function isAlianzaRow(r){return isAlianza(companyText(r));}
function reportMonths(){return [DATA.months.at(-2)?.id,DATA.months.at(-1)?.id].filter(Boolean);}
function downloadStats(report){
  const [base,comp]=reportMonths(),br=(DATA.details[base]||[]).filter(report.filter),cr=(DATA.details[comp]||[]).filter(report.filter);
  return {base,comp,baseCount:br.length,compCount:cr.length};
}
function renderDownloads(){
  const [base,comp]=reportMonths();
  downloadSub.textContent=`${label(base)} → ${label(comp)} · fuente: ${DATA.metadata.source}`;
  const total=downloadReports.reduce((a,r)=>{const s=downloadStats(r);a.base+=s.baseCount;a.comp+=s.compCount;return a;},{base:0,comp:0});
  downloadGrid.innerHTML=`<div class="rc" style="border-top:3px solid var(--blue)"><div class="rv" style="color:var(--blue)">${downloadReports.length}</div><div class="rl">Archivos</div><div class="rs">Cortes configurados</div></div><div class="rc" style="border-top:3px solid var(--green)"><div class="rv" style="color:var(--green)">${label(base)}</div><div class="rl">Mes anterior</div><div class="rs">${N(total.base)} registros en cortes</div></div><div class="rc" style="border-top:3px solid var(--purple)"><div class="rv" style="color:var(--purple)">${label(comp)}</div><div class="rl">Mes actual</div><div class="rs">${N(total.comp)} registros en cortes</div></div>`;
  downloadBody.innerHTML=downloadReports.map(r=>{const s=downloadStats(r);return `<tr><td class="tn">${r.name}</td><td>${r.rule}</td><td class="nr">${N(s.baseCount)}</td><td class="nr">${N(s.compCount)}</td><td class="nr"><button class="btn" onclick="downloadRevision('${r.id}')">Descargar</button></td></tr>`}).join('');
}
function summarizeDownloadRows(rows,groupKey){
  const map={};
  rows.forEach(r=>{const k=r[groupKey]||'Sin dato';const x=map[k]||(map[k]={grupo:k,dotacion:0,fte:0,liquido:0,haberes:0});x.dotacion++;x.fte+=(+r.fte||(+r.dias||0)/30);x.liquido+=(+r.sueldo_liquido||0);x.haberes+=(+r.total_haberes||0);});
  return map;
}
function summarySheetAoA(title,baseRows,compRows,groupKey,base,comp){
  const bm=summarizeDownloadRows(baseRows,groupKey),cm=summarizeDownloadRows(compRows,groupKey),keys=[...new Set([...Object.keys(bm),...Object.keys(cm)])].sort();
  const rows=[['',title,'Dotación',null,null,'FTE',null,null,'Sueldo Líquido',null,null,'Total Haberes',null,null],['','',base,comp,'Diferencia',base,comp,'Diferencia',base,comp,'Diferencia',base,comp,'Diferencia']];
  let total={bd:0,cd:0,bf:0,cf:0,bl:0,cl:0,bh:0,ch:0};
  keys.forEach((k,i)=>{const b=bm[k]||{},c=cm[k]||{};total.bd+=b.dotacion||0;total.cd+=c.dotacion||0;total.bf+=b.fte||0;total.cf+=c.fte||0;total.bl+=b.liquido||0;total.cl+=c.liquido||0;total.bh+=b.haberes||0;total.ch+=c.haberes||0;rows.push([i+1,k,b.dotacion||0,c.dotacion||0,(c.dotacion||0)-(b.dotacion||0),+(b.fte||0).toFixed(2),+(c.fte||0).toFixed(2),+((c.fte||0)-(b.fte||0)).toFixed(2),Math.round(b.liquido||0),Math.round(c.liquido||0),Math.round((c.liquido||0)-(b.liquido||0)),Math.round(b.haberes||0),Math.round(c.haberes||0),Math.round((c.haberes||0)-(b.haberes||0))]);});
  rows.push(['','TOTAL',total.bd,total.cd,total.cd-total.bd,+total.bf.toFixed(2),+total.cf.toFixed(2),+(total.cf-total.bf).toFixed(2),Math.round(total.bl),Math.round(total.cl),Math.round(total.cl-total.bl),Math.round(total.bh),Math.round(total.ch),Math.round(total.ch-total.bh)]);
  return rows;
}
function detailSheetAoA(baseRows,compRows,base,comp){
  const bm=Object.fromEntries(baseRows.map(r=>[r.rut,r])),cm=Object.fromEntries(compRows.map(r=>[r.rut,r])),keys=[...new Set([...Object.keys(bm),...Object.keys(cm)])].sort((a,b)=>String((cm[a]||bm[a]).empresa).localeCompare(String((cm[b]||bm[b]).empresa))||String((cm[a]||bm[a]).sede).localeCompare(String((cm[b]||bm[b]).sede))||String((cm[a]||bm[a]).nombre).localeCompare(String((cm[b]||bm[b]).nombre)));
  const rows=[['N°','Empresa','Nombre','Rut','Sede','Cargo','Días Trab.',null,null,'Sueldo Líquido',null,null,'Total Haberes',null,null,'Sueldo Base',null,null],['','','','','','',base,comp,'Diferencia',base,comp,'Diferencia',base,comp,'Diferencia',base,comp,'Diferencia']];
  keys.forEach((rut,i)=>{const b=bm[rut]||{},c=cm[rut]||{},r=c.rut?c:b;rows.push([i+1,r.empresa,r.nombre,r.rut,r.sede,r.cargo,b.dias||0,c.dias||0,(c.dias||0)-(b.dias||0),b.sueldo_liquido||0,c.sueldo_liquido||0,(c.sueldo_liquido||0)-(b.sueldo_liquido||0),b.total_haberes||0,c.total_haberes||0,(c.total_haberes||0)-(b.total_haberes||0),b.sueldo_base||0,c.sueldo_base||0,(c.sueldo_base||0)-(b.sueldo_base||0)]);});
  return rows;
}
function monthRowsAoA(rows,month){
  const headers=DATA.metadata.raw_headers?.length?DATA.metadata.raw_headers:['Empresa','Nombre empresa','Rut empresa','Proceso','Nombre','Rut','Contrato','Sede','Días Trabajados','Cargo'];
  return [headers,...rows.map(r=>{
    if(r.raw?.length) return headers.map((_,i)=>r.raw[i]??'');
    const fallback=['',r.empresa,'',month,r.nombre,r.rut,r.contrato,r.sede,r.dias,r.cargo];
    return headers.map((_,i)=>fallback[i]??'');
  })];
}
const xlStyles={
  top:{fill:{patternType:'solid',fgColor:{rgb:'1F3864'}},font:{name:'Calibri',sz:10,bold:true,color:{rgb:'FFFFFF'}},alignment:{horizontal:'center',vertical:'center'},border:{bottom:{style:'thin',color:{rgb:'D9E2F3'}}}},
  sub:{fill:{patternType:'solid',fgColor:{rgb:'2E75B6'}},font:{name:'Calibri',sz:10,bold:true,color:{rgb:'FFFFFF'}},alignment:{horizontal:'center',vertical:'center'},border:{bottom:{style:'thin',color:{rgb:'D9E2F3'}}}},
  body:{font:{name:'Calibri',sz:10,color:{rgb:'000000'}},alignment:{vertical:'center'},border:{bottom:{style:'thin',color:{rgb:'E7EAF0'}}}},
  body9:{font:{name:'Calibri',sz:9,color:{rgb:'000000'}},alignment:{vertical:'center'},border:{bottom:{style:'thin',color:{rgb:'E7EAF0'}}}},
  zebra:{fill:{patternType:'solid',fgColor:{rgb:'F2F2F2'}},font:{name:'Calibri',sz:9,color:{rgb:'000000'}},alignment:{vertical:'center'},border:{bottom:{style:'thin',color:{rgb:'E7EAF0'}}}},
  total:{fill:{patternType:'solid',fgColor:{rgb:'D9EAF7'}},font:{name:'Calibri',sz:10,bold:true,color:{rgb:'000000'}},alignment:{vertical:'center'},border:{top:{style:'thin',color:{rgb:'1F3864'}},bottom:{style:'thin',color:{rgb:'1F3864'}}}},
  diffPos:{font:{name:'Calibri',sz:10,color:{rgb:'008000'}},alignment:{vertical:'center'}},
  diffNeg:{font:{name:'Calibri',sz:10,color:{rgb:'C00000'}},alignment:{vertical:'center'}},
};
function encodeCell(r,c){return XLSX.utils.encode_cell({r,c});}
function rangeRef(rows,cols){return XLSX.utils.encode_range({s:{r:0,c:0},e:{r:Math.max(rows-1,0),c:Math.max(cols-1,0)}});}
function styleSheet(ws,aoa,kind){
  const rows=aoa.length,cols=Math.max(...aoa.map(r=>r.length),1);
  for(let r=0;r<rows;r++){
    for(let c=0;c<cols;c++){
      const ref=encodeCell(r,c),cell=ws[ref]; if(!cell) continue;
      if(kind==='month') cell.s=r===0?xlStyles.top:(r%2?xlStyles.body9:xlStyles.zebra);
      else if(kind==='detail') cell.s=r<=1?xlStyles.top:xlStyles.body9;
      else cell.s=r===0?xlStyles.top:r===1?xlStyles.sub:r===rows-1?xlStyles.total:xlStyles.body;
      if(typeof cell.v==='number'){
        cell.z=c>=6?'#,##0;[Red](#,##0);-':'#,##0';
        cell.s={...cell.s,alignment:{...(cell.s.alignment||{}),horizontal:'right'}};
      }
      if((kind==='detail'||kind==='summary')&&r>=2&&[4,7,10,13,17].includes(c)&&typeof cell.v==='number'){
        cell.s={...cell.s,...(cell.v<0?xlStyles.diffNeg:cell.v>0?xlStyles.diffPos:{})};
        cell.z='#,##0;[Red](#,##0);-';
      }
    }
  }
  ws['!autofilter']={ref:kind==='summary'?`A2:${XLSX.utils.encode_col(cols-1)}${rows}`:`A1:${XLSX.utils.encode_col(cols-1)}${rows}`};
  ws['!cols']=Array.from({length:cols},(_,i)=>({wch:kind==='detail'?[5,28,34,13,18,28,9,12,12,14,14,14,14,14,14,14,14,14][i]||13:kind==='summary'?[4,32,9,13,11,9,13,11,14,14,14,14,14,14][i]||13:13}));
}
function mergeRange(sRow,sCol,eRow,eCol){return {s:{r:sRow,c:sCol},e:{r:eRow,c:eCol}};}
function headerMerges(kind){
  if(kind==='summary') return [
    mergeRange(0,0,1,0),mergeRange(0,1,1,1),
    mergeRange(0,2,0,4),mergeRange(0,5,0,7),mergeRange(0,8,0,10),mergeRange(0,11,0,13),
  ];
  if(kind==='detail') return [
    mergeRange(0,0,1,0),mergeRange(0,1,1,1),mergeRange(0,2,1,2),mergeRange(0,3,1,3),mergeRange(0,4,1,4),mergeRange(0,5,1,5),
    mergeRange(0,6,0,8),mergeRange(0,9,0,11),mergeRange(0,12,0,14),mergeRange(0,15,0,17),
  ];
  return [];
}
function appendSheet(wb,name,aoa,kind='summary'){
  const ws=XLSX.utils.aoa_to_sheet(aoa);
  styleSheet(ws,aoa,kind);
  const merges=headerMerges(kind);
  if(merges.length) ws['!merges']=merges;
  if(kind==='summary'){ws['!freeze']={xSplit:2,ySplit:2};ws['!views']=[{state:'frozen',xSplit:2,ySplit:2,topLeftCell:'C3'}];}
  if(kind==='detail'){ws['!freeze']={xSplit:6,ySplit:2};ws['!views']=[{state:'frozen',xSplit:6,ySplit:2,topLeftCell:'G3'}];}
  if(kind==='month'){ws['!freeze']={xSplit:0,ySplit:1};ws['!views']=[{state:'frozen',ySplit:1,topLeftCell:'A2'}];}
  XLSX.utils.book_append_sheet(wb,ws,name);
}
function downloadRevision(id){
  if(!window.XLSX){alert('No se pudo cargar la librería XLSX. Revisa la conexión a internet.');return;}
  const report=downloadReports.find(r=>r.id===id); if(!report) return;
  const [base,comp]=reportMonths(); if(!base||!comp){alert('Se necesitan al menos 2 meses en el archivo cargado.');return;}
  const baseRows=(DATA.details[base]||[]).filter(report.filter),compRows=(DATA.details[comp]||[]).filter(report.filter);
  const wb=XLSX.utils.book_new();
  appendSheet(wb,'Resumen por empresa',summarySheetAoA('Grupo',baseRows,compRows,'empresa',base,comp),'summary');
  appendSheet(wb,'Resumen completo',summarySheetAoA('Grupo',baseRows,compRows,'sede',base,comp),'summary');
  appendSheet(wb,'Detalle individual',detailSheetAoA(baseRows,compRows,base,comp),'detail');
  appendSheet(wb,'Mes anterior',monthRowsAoA(baseRows,base),'month');
  appendSheet(wb,'Mes actual',monthRowsAoA(compRows,comp),'month');
  XLSX.writeFile(wb,`${report.file}_${String(comp).replace('-','_')}.xlsx`);
}
function activateTab(sec){
  document.querySelectorAll('.tab,.sec').forEach(x=>x.classList.remove('active'));
  document.querySelector(`.tab[data-sec="${sec}"]`)?.classList.add('active');
  document.getElementById('sec-'+sec)?.classList.add('active');
  if(sec==='detalle') setTimeout(initDetailScrollSync,0);
}
function goDetail(month,empresa='',sede=''){
  detMonth.value=month||state.comp;
  detEmpresa.value=empresa||'';
  srch.value='';
  detCargo.value='';
  detConcept.value='';
  detDias.value='';
  refreshDetailFilters();
  detEmpresa.value=empresa||'';
  detSede.value=sede||'';
  renderDetail();
  activateTab('detalle');
}
function goWorkerDetail(month,rut){
  detMonth.value=month||state.comp;
  detEmpresa.value='';
  detSede.value='';
  detCargo.value='';
  detConcept.value='';
  detDias.value='';
  srch.value=rut||'';
  refreshDetailFilters();
  renderDetail();
  activateTab('detalle');
}
function refreshDetailFilters(){
  const month=detMonth.value || state.comp, rows=DATA.details[month]||[],currentEmpresa=detEmpresa.value,currentSede=detSede.value,currentCargo=detCargo.value,currentConcept=detConcept.value;
  const empresas=[...new Set(rows.map(r=>r.empresa))].sort();
  detEmpresa.innerHTML='<option value="">Todas las empresas</option>'+empresas.map(e=>`<option value="${e}">${e}</option>`).join('');
  if(empresas.includes(currentEmpresa)) detEmpresa.value=currentEmpresa;
  const filtered=detEmpresa.value?rows.filter(r=>r.empresa===detEmpresa.value):rows;
  const sedes=[...new Set(filtered.map(r=>r.sede))].sort();
  detSede.innerHTML='<option value="">Todas las sedes</option>'+sedes.map(s=>`<option value="${s}">${s}</option>`).join('');
  if(sedes.includes(currentSede)) detSede.value=currentSede;
  const filtered2=filtered.filter(r=>!detSede.value||r.sede===detSede.value);
  const cargos=[...new Set(filtered2.map(r=>r.cargo).filter(Boolean))].sort();
  detCargo.innerHTML='<option value="">Todos los cargos</option>'+cargos.map(c=>`<option value="${c}">${c}</option>`).join('');
  if(cargos.includes(currentCargo)) detCargo.value=currentCargo;
  detConcept.innerHTML='<option value="">Todos los conceptos</option>'+DATA.concept_options.map(c=>`<option value="${c}">${c}</option>`).join('');
  if(DATA.concept_options.includes(currentConcept)) detConcept.value=currentConcept;
}
function detailColumns(){
  const fixed=[
    {key:'empresa',label:'Empresa',cls:'fix f1',type:'text'},
    {key:'sede',label:'Sede',cls:'fix f2',type:'text'},
    {key:'nombre',label:'Nombre',cls:'fix f3',type:'text'},
  ];
  const baseCols=[
    {key:'rut',label:'RUT',type:'text'},
    {key:'contrato',label:'Contrato',type:'text'},
    {key:'cargo',label:'Cargo',type:'text'},
    {key:'dias',label:'Días',type:'dias'},
    {key:'fte',label:'FTE',type:'number'},
    {key:'ausentismo_dias',label:'Días Ausentismo',type:'number'},
    {key:'licencia_dias',label:'Días Licencia',type:'number'},
  ];
  const conceptCols=DATA.concept_options.map(c=>({key:`concept::${c}`,label:c,type:'money'}));
  const totalCols=[
    {key:'total_haberes',label:'Suma Haberes',type:'money'},
    {key:'total_descuentos',label:'Total Rebajas',type:'money'},
    {key:'sueldo_liquido',label:'Sueldo Líquido',type:'money'},
  ];
  return [...fixed,...baseCols,...conceptCols,...totalCols];
}
function detailRawValue(r,c){
  if(c.key.startsWith('concept::')) return r.concepts?.[c.key.replace('concept::','')]||0;
  return r[c.key] ?? '';
}
function detailDisplayValue(r,c){
  const raw=detailRawValue(r,c);
  if(c.type==='money') return raw?K(raw):'-';
  if(c.type==='number') return (+raw||0).toFixed(2);
  if(c.type==='dias') return String(+raw||0);
  return String(raw||'');
}
function detailRowPassesColumnFilters(r,cols,skipKey=''){
  return cols.every(c=>{
    if(c.key===skipKey) return true;
    const selected=state.detailFilters?.[c.key]||[];
    return !selected.length || selected.includes(detailDisplayValue(r,c));
  });
}
function detailGlobalRows(){
  const month=detMonth.value || state.comp,q=srch.value.toLowerCase(),empresa=detEmpresa.value,sede=detSede.value,dias=detDias.value,cargo=detCargo.value,concept=detConcept.value;
  return (DATA.details[month]||[]).filter(r=>{
    if(empresa&&r.empresa!==empresa)return false; if(sede&&r.sede!==sede)return false;
    if(cargo&&r.cargo!==cargo)return false; if(concept&&!(r.concepts?.[concept]))return false;
    if(dias==='0'&&r.dias!==0)return false; if(dias==='p'&&!(r.dias>0&&r.dias<30))return false; if(dias==='f'&&r.dias<30)return false;
    if(q&&!`${r.rut} ${r.nombre} ${r.cargo} ${r.empresa} ${r.sede}`.toLowerCase().includes(q))return false;
    return true;
  });
}
function openDetailFilterMenu(ev,key){
  ev.stopPropagation();
  detailFilterKey=key;
  const cols=detailColumns(), col=cols.find(c=>c.key===key);
  if(!col) return;
  const current=new Set(state.detailFilters?.[key]||[]);
  const rows=detailGlobalRows().filter(r=>detailRowPassesColumnFilters(r,cols,key));
  const values=[...new Set(rows.map(r=>detailDisplayValue(r,col)))].sort((a,b)=>String(a).localeCompare(String(b),'es',{numeric:true,sensitivity:'base'}));
  const visible=values;
  const menu=detailFilterMenu;
  menu.innerHTML=`<div class="df-title">${txt(col.label)}</div><input class="df-search" placeholder="Buscar opciones..." oninput="filterDetailMenuOptions(this.value)"><div class="df-actions"><button onclick="selectAllDetailFilter(true)">Seleccionar todo</button><button onclick="selectAllDetailFilter(false)">Limpiar</button></div><div class="df-list" data-total="${values.length}">${visible.map(v=>`<label class="df-opt"><input type="checkbox" value="${txt(v)}" ${(!current.size||current.has(v))?'checked':''}><span>${txt(v)}</span></label>`).join('')}</div><div class="df-foot"><button class="df-apply" onclick="applyDetailFilter()">Aplicar</button></div>`;
  const rect=ev.currentTarget.getBoundingClientRect();
  menu.style.left=Math.min(rect.left,window.innerWidth-318)+'px';
  menu.style.top=Math.min(rect.bottom+6,window.innerHeight-410)+'px';
  menu.classList.add('open');
}
function filterDetailMenuOptions(q){
  q=String(q||'').toLowerCase();
  detailFilterMenu.querySelectorAll('.df-opt').forEach(l=>{
    l.style.display=l.textContent.toLowerCase().includes(q)?'flex':'none';
  });
}
function selectAllDetailFilter(flag){
  detailFilterMenu.querySelectorAll('.df-opt').forEach(l=>{
    if(l.style.display!=='none') l.querySelector('input').checked=flag;
  });
}
function applyDetailFilter(){
  const checked=[...detailFilterMenu.querySelectorAll('.df-opt input:checked')].map(i=>i.value);
  const total=+(detailFilterMenu.querySelector('.df-list')?.dataset.total||0);
  state.detailFilters=state.detailFilters||{};
  if(!checked.length || checked.length===total) delete state.detailFilters[detailFilterKey];
  else state.detailFilters[detailFilterKey]=checked;
  closeDetailFilterMenu();
  renderDetail();
}
function closeDetailFilterMenu(){detailFilterMenu?.classList.remove('open')}
function initDetailScrollSync(){
  const top=document.getElementById('detailTopScroll'),inner=document.getElementById('detailTopInner'),wrap=document.getElementById('detailBookWrap');
  if(!top||!inner||!wrap) return;
  inner.style.width=wrap.scrollWidth+'px';
  if(top.dataset.bound) return;
  let lock=false;
  top.addEventListener('scroll',()=>{if(lock)return;lock=true;wrap.scrollLeft=top.scrollLeft;lock=false;});
  wrap.addEventListener('scroll',()=>{if(lock)return;lock=true;top.scrollLeft=wrap.scrollLeft;lock=false;});
  top.dataset.bound='1';
}
function renderDetail(){
  const cols=detailColumns();
  closeDetailFilterMenu();
  let rows=detailGlobalRows().filter(r=>detailRowPassesColumnFilters(r,cols)).sort((a,b)=>b.total_haberes-a.total_haberes);
  const activeFilters=Object.values(state.detailFilters||{}).filter(v=>v?.length).length;
  detCnt.textContent=`${N(rows.length)} registros · ${N(cols.length)} columnas${activeFilters?` · ${activeFilters} filtros de cabecera`:''}`;
  const colClass=c=>[c.cls||'',c.type==='money'?'nr money-col':c.type==='number'?'nr':c.type==='text'?'text-col':''].join(' ').trim();
  detHead.innerHTML=cols.map(c=>{const selected=state.detailFilters?.[c.key]||[];return `<th class="${colClass(c)}"><span class="th-label">${txt(c.label)}</span><button class="head-filter-btn ${selected.length?'active':''}" onclick="openDetailFilterMenu(event,'${esc(c.key)}')"><span>${selected.length?`${selected.length} sel.`:'Todos'}</span><b>▾</b></button></th>`}).join('');
  function cellValue(r,c){
    if(c.type==='dias'){
      const v=+r.dias||0;
      return v===0?`<span class="bdg br2">${v}</span>`:v<30?`<span class="bdg by">${v}</span>`:`<span class="bdg bg">${v}</span>`;
    }
    if(c.type==='money'){
      const raw=detailRawValue(r,c);
      return raw?K(raw):'<span class="zero-cell">-</span>';
    }
    if(c.type==='number') return `<span style="font-family:var(--m)">${(+r[c.key]||0).toFixed(2)}</span>`;
    if(c.key==='sede') return `<span class="bdg bb">${txt(r.sede)}</span>`;
    if(c.key==='rut') return `<span class="muted-cell" style="font-family:var(--m)">${txt(r.rut)}</span>`;
    if(c.key==='cargo'||c.key==='contrato') return `<span class="muted-cell">${txt(r[c.key])}</span>`;
    return txt(r[c.key]);
  }
  detBody.innerHTML=rows.map(r=>`<tr>${cols.map(c=>`<td class="${colClass(c)}">${cellValue(r,c)}</td>`).join('')}</tr>`).join('');
  initDetailScrollSync();
}
function parseNumber(value){
  if(value===null||value===undefined||value==='') return 0;
  if(typeof value==='number') return Number.isFinite(value)?value:0;
  const s=String(value).replace(/\\./g,'').replace(',','.');
  const n=Number(s);
  return Number.isFinite(n)?n:0;
}
function cleanText(value){return String(value??'').trim();}
function headerKey(value){return cleanText(value).normalize('NFD').replace(/[\u0300-\u036f]/g,'').toLowerCase();}
function headerIndex(headers){
  const out={};
  headers.forEach((h,i)=>{const raw=cleanText(h),key=headerKey(h);if(raw&&!out[raw])out[raw]=i;if(key&&!out[key])out[key]=i;});
  return out;
}
function idxOf(idx,name){return idx[name]??idx[headerKey(name)];}
function rawVal(raw,idx,name){const i=idxOf(idx,name);return i===undefined?'':raw[i];}
function numVal(raw,idx,name){return parseNumber(rawVal(raw,idx,name));}
function sumHeaders(raw,headers,pred){return headers.reduce((a,h,i)=>a+(pred(headerKey(h),cleanText(h))?parseNumber(raw[i]||0):0),0);}
function monthLabelFromPeriod(period){
  const names={'01':'Ene','02':'Feb','03':'Mar','04':'Abr','05':'May','06':'Jun','07':'Jul','08':'Ago','09':'Sep','10':'Oct','11':'Nov','12':'Dic'};
  const [y,m]=String(period).split('-');
  return `${names[m]||m} ${y}`;
}
function summarizeRows(rows){
  const dot=rows.length,dias=rows.reduce((a,r)=>a+r.dias,0),sum=k=>rows.reduce((a,r)=>a+(+r[k]||0),0),fte=rows.reduce((a,r)=>a+(+r.dias||0)/30,0);
  const cero=rows.filter(r=>r.dias===0).length,menos=rows.filter(r=>r.dias>0&&r.dias<30).length;
  return {
    dotacion:dot,fte:+fte.toFixed(2),dias_prom:dot?+(dias/dot).toFixed(1):0,cero_dias:cero,menos30:menos,
    rotacion_pct:dot?+((cero+menos)/dot*100).toFixed(1):0,ausentes_pct:dot?+(cero/dot*100).toFixed(1):0,
    total_haberes:Math.round(sum('total_haberes')),sueldo_base:Math.round(sum('sueldo_base')),sueldo_liquido:Math.round(sum('sueldo_liquido')),total_descuentos:Math.round(sum('total_descuentos')),
    licencia_dias:+sum('licencia_dias').toFixed(1),ausentismo_dias:+sum('ausentismo_dias').toFixed(1),licencia_personas:rows.filter(r=>(+r.licencia_dias||0)>0).length,ausentismo_personas:rows.filter(r=>(+r.ausentismo_dias||0)>0).length,
    gratificacion:Math.round(sum('gratificacion')),hhee:Math.round(sum('hhee')),movilizacion:Math.round(sum('movilizacion')),colacion:Math.round(sum('colacion')),
    sb_promedio:dot?Math.round(sum('sueldo_base')/dot):0,liq_promedio:dot?Math.round(sum('sueldo_liquido')/dot):0,pct_fte:dot?+(fte/dot*100).toFixed(1):0
  };
}
function groupRowsJs(rows,keys){
  const buckets={};
  rows.forEach(r=>{const k=keys.map(x=>r[x]||'Sin dato').join(' · ');(buckets[k]=buckets[k]||[]).push(r);});
  return Object.entries(buckets).map(([key,items])=>{const out=summarizeRows(items);keys.forEach((k,i)=>out[k]=items[0][k]||'Sin dato');out.key=key;return out;}).sort((a,b)=>b.total_haberes-a.total_haberes);
}
function groupConceptsJs(rows,conceptCols){
  const grouped={total:{},empresa:{},sede:{},empresa_sede:{}};
  const add=(bucket,key,concept,value)=>{const obj=key?((grouped[bucket][key]=grouped[bucket][key]||{})):grouped[bucket];obj[concept]=(obj[concept]||0)+value;};
  rows.forEach(r=>{
    const emp=r.empresa,sede=r.sede,empSede=`${emp} · ${sede}`;
    conceptCols.forEach(({idx,label})=>{
      const value=parseNumber(r._raw[idx]||0);
      add('total','',label,value);add('empresa',emp,label,value);add('sede',sede,label,value);add('empresa_sede',empSede,label,value);
    });
  });
  return grouped;
}
function buildDataFromRows(aoa,fileName){
  const headerRow=aoa.findIndex(row=>row&&headerKey(row[0])==='empresa'&&headerKey(row[3])==='proceso');
  const headers=aoa[headerRow>=0?headerRow:4]||[],idx=headerIndex(headers);
  const need=['Proceso','Nombre empresa','Sede','Nombre','Rut','Contrato','Cargo','Días Trabajados','Sueldo Base','Suma Haberes','Sueldo Líquido','Total Rebajas'];
  const missing=need.filter(h=>idxOf(idx,h)===undefined);
  if(missing.length) throw new Error(`Faltan columnas: ${missing.join(', ')}`);
  const conceptCols=[];
  const conceptStart=idxOf(idx,'Sueldo Base')??22,conceptEnd=[idxOf(idx,'Cotizacion AFP'),idxOf(idx,'Sueldo Líquido')].filter(x=>x!==undefined).sort((a,b)=>a-b)[0]??headers.length;
  for(let i=conceptStart;i<conceptEnd;i++){const label=cleanText(headers[i]);if(label&&headerKey(label)!=='suma haberes'&&headerKey(label)!=='sueldo liquido') conceptCols.push({idx:i,label,type:'haber'});}
  const discountStart=idxOf(idx,'Cotizacion AFP'),discountEnd=[idxOf(idx,'Aporte a CCAF'),idxOf(idx,'Mutual'),idxOf(idx,'Sueldo Líquido')].filter(x=>x!==undefined).sort((a,b)=>a-b)[0]??headers.length;
  if(discountStart!==undefined){for(let i=discountStart;i<discountEnd;i++){const label=cleanText(headers[i]);if(label&&headerKey(label)!=='total rebajas') conceptCols.push({idx:i,label,type:'descuento'});}}
  const byMonth={},conceptSums={},companies=new Set();
  aoa.slice((headerRow>=0?headerRow:4)+1).forEach(raw=>{
    if(!raw||!raw[0]||raw[0]==='Empresa') return;
    const period=cleanText(rawVal(raw,idx,'Proceso')); if(!period) return;
    const empresa=cleanText(rawVal(raw,idx,'Nombre empresa')),sede=cleanText(rawVal(raw,idx,'Sede'))||'Sin sede';
    const conceptValues={};
    conceptCols.forEach(c=>{
      const v=parseNumber(raw[c.idx]||0);
      if(v!==0) conceptValues[c.label]=v;
      const monthSums=conceptSums[period]=conceptSums[period]||{};
      monthSums[c.label]=(monthSums[c.label]||0)+v;
    });
    const row={_raw:headers.map((_,i)=>raw[i]??''),proceso:period,empresa,sede,nombre:cleanText(rawVal(raw,idx,'Nombre')),rut:cleanText(rawVal(raw,idx,'Rut')),contrato:cleanText(rawVal(raw,idx,'Contrato')),cargo:cleanText(rawVal(raw,idx,'Cargo')),
      dias:numVal(raw,idx,'Días Trabajados'),sueldo_base:numVal(raw,idx,'Sueldo Base'),total_haberes:numVal(raw,idx,'Suma Haberes'),sueldo_liquido:numVal(raw,idx,'Sueldo Líquido'),
      total_descuentos:numVal(raw,idx,'Total Rebajas'),gratificacion:numVal(raw,idx,'Gratificación'),hhee:numVal(raw,idx,'Horas Extras Empresa 50%'),
      licencia_dias:sumHeaders(raw,headers,k=>k.includes('licencia')),ausentismo_dias:sumHeaders(raw,headers,k=>k.includes('licencia')||k.includes('permiso')||k.includes('falta')),
      movilizacion:numVal(raw,idx,'Movilizacion'),colacion:numVal(raw,idx,'Colacion'),
      bono_manip_pae:numVal(raw,idx,'Bono Manipuladora Pae')+numVal(raw,idx,'Bono Manipuladora Pae I'),concept_values:conceptValues};
    (byMonth[period]=byMonth[period]||[]).push(row); companies.add(empresa);
  });
  const conceptTypes={};
  conceptCols.forEach(c=>{conceptTypes[c.label]=c.type;});
  const months=Object.keys(byMonth).sort(),data={metadata:{source:fileName,generated_from:'Detalle',month_count:months.length,record_count:Object.values(byMonth).reduce((a,r)=>a+r.length,0),company_count:companies.size,raw_headers:headers},months:months.map(m=>({id:m,label:monthLabelFromPeriod(m)})),kpis:{},groups:{empresa:{},sede:{},empresa_sede:{}},concepts:{},concept_groups:{},concept_options:[...new Set(conceptCols.map(c=>c.label))].sort(),concept_types:conceptTypes,details:{}};
  months.forEach(month=>{
    const rows=byMonth[month];
    data.kpis[month]=summarizeRows(rows);
    data.groups.empresa[month]=groupRowsJs(rows,['empresa']);
    data.groups.sede[month]=groupRowsJs(rows,['sede']);
    data.groups.empresa_sede[month]=groupRowsJs(rows,['empresa','sede']);
    data.concepts[month]=conceptSums[month]||{};
    data.concept_groups[month]=groupConceptsJs(rows,conceptCols);
    data.details[month]=rows.map(r=>({rut:r.rut,nombre:r.nombre,cargo:r.cargo,contrato:r.contrato,empresa:r.empresa,sede:r.sede,dias:r.dias,fte:+(r.dias/30).toFixed(2),sueldo_base:Math.round(r.sueldo_base),total_haberes:Math.round(r.total_haberes),total_descuentos:Math.round(r.total_descuentos),sueldo_liquido:Math.round(r.sueldo_liquido),hhee:Math.round(r.hhee),licencia_dias:+(r.licencia_dias||0).toFixed(1),ausentismo_dias:+(r.ausentismo_dias||0).toFixed(1),bono_manip_pae:Math.round(r.bono_manip_pae),raw:r._raw,concepts:Object.fromEntries(Object.entries(r.concept_values).map(([k,v])=>[k,Math.round(v)]))}));
  });
  return data;
}
function reloadData(newData){
  DATA=newData; byId=Object.fromEntries(DATA.months.map(m=>[m.id,m]));
  state.base=DATA.months.at(-2)?.id||DATA.months[0]?.id||''; state.comp=DATA.months.at(-1)?.id||state.base;
    state.empresa='';state.sede='';state.metric='total_haberes';state.compareMetrics=['total_haberes','sueldo_liquido'];state.expandedCompanies={};state.expandedKpiRot={};state.expandedKpiAus={};state.expandedKpiLic={};state.expandedDiffConcepts={};state.expandedDiffCompanies={};state.expandedDiffSedes={};state.selectedConcept='';state.detailFilters={};
  initSelectors(); renderAll();
}
function handleFileUpload(e){
  const file=e.target.files?.[0]; if(!file) return;
  if(!window.XLSX){alert('No se pudo cargar la librería XLSX. Revisa la conexión a internet.');return;}
  const reader=new FileReader();
  reader.onload=ev=>{
    try{
      const wb=XLSX.read(new Uint8Array(ev.target.result),{type:'array',cellDates:false});
      const ws=wb.Sheets.Detalle||wb.Sheets[wb.SheetNames[0]];
      const aoa=XLSX.utils.sheet_to_json(ws,{header:1,defval:'',raw:false});
      reloadData(buildDataFromRows(aoa,file.name));
    }catch(err){console.error(err);alert(`No pude cargar el archivo: ${err.message}`);}
  };
  reader.readAsArrayBuffer(file);
}
function renderAll(){renderChips();renderCompanyDivision();renderKPIs();renderTotalsHighlights();renderKpisMenu();renderComparativo();renderConcepts();renderDeviaciones();renderAuditoria();renderDownloads();renderDetail();}
document.querySelectorAll('.tab').forEach(t=>t.addEventListener('click',()=>activateTab(t.dataset.sec)));
document.addEventListener('click',e=>{if(!e.target.closest?.('#detailFilterMenu')&&!e.target.closest?.('.head-filter-btn'))closeDetailFilterMenu();});
initSelectors(); renderAll();
</script>
</body>
</html>
"""


def main() -> None:
    data = build_data()
    html = HTML_TEMPLATE.replace("__DATA__", json.dumps(data, ensure_ascii=False, separators=(",", ":")))
    OUTPUT.write_text(html, encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(json.dumps(data["metadata"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
